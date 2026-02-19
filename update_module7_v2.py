"""
Update Module7.xlsm with Lululemon Athletica (LULU) Data
========================================================
- Uses actual year labels (2021, 2022, 2023, 2024, 2025, 2026E, 2027E)
- Implements Assets = Liabilities + Equity formulas
"""

import openpyxl
import openpyxl.cell.cell
from openpyxl.styles import Font

def update_module7():
    # Load the template (preserve macros)
    filepath = r"c:\Users\nduta\OneDrive\Desktop\Projects\lulu-lemon-project\Module7.xlsm"
    wb = openpyxl.load_workbook(filepath, keep_vba=True)
    
    # =========================================================================
    # LULULEMON DATA FROM BLOOMBERG ($ in millions)
    # =========================================================================
    
    # Income Statement Data
    income_data = {
        '2021': {
            'Revenue': 4401.879,
            'COGS': 1937.888,
            'Gross_Profit': 2463.991,
            'SGA': 1609.003,
            'Other_OpEx': 35.002,
            'EBITDA': 1005.464,
            'DA': 185.478,
            'EBIT': 819.986,
            'NonOp_Income': -0.636,
            'Interest_Expense': 0,
            'Unusual_Expense': 0,
            'Pretax_Income': 819.35,
            'Income_Tax': 230.437,
            'Net_Income': 588.913,
            'Pref_Dividends': 0,
            'Net_Income_Common': 588.913,
            'Basic_Shares': 130.5,
            'Diluted_Shares': 130.9,
        },
        '2022': {
            'Revenue': 6256.617,
            'COGS': 2648.052,
            'Gross_Profit': 3608.565,
            'SGA': 2225.034,
            'Other_OpEx': 50.176,
            'EBITDA': 1557.561,
            'DA': 224.206,
            'EBIT': 1333.355,
            'NonOp_Income': 0.514,
            'Interest_Expense': 0,
            'Unusual_Expense': 0,
            'Pretax_Income': 1333.869,
            'Income_Tax': 358.547,
            'Net_Income': 975.322,
            'Pref_Dividends': 0,
            'Net_Income_Common': 975.322,
            'Basic_Shares': 129.5,
            'Diluted_Shares': 130.2,
        },
        '2023': {
            'Revenue': 8110.518,
            'COGS': 3618.178,
            'Gross_Profit': 4492.34,
            'SGA': 2757.447,
            'Other_OpEx': 406.485,
            'EBITDA': 1620.199,
            'DA': 291.791,
            'EBIT': 1328.408,
            'NonOp_Income': 401.896,
            'Interest_Expense': 0,
            'Unusual_Expense': 0,
            'Pretax_Income': 1730.304,
            'Income_Tax': 875.504,
            'Net_Income': 854.8,
            'Pref_Dividends': 0,
            'Net_Income_Common': 854.8,
            'Basic_Shares': 127.2,
            'Diluted_Shares': 128.0,
        },
        '2024': {
            'Revenue': 9619.278,
            'COGS': 4009.873,
            'Gross_Profit': 5609.405,
            'SGA': 3397.218,
            'Other_OpEx': 79.511,
            'EBITDA': 2512.06,
            'DA': 379.384,
            'EBIT': 2132.676,
            'NonOp_Income': 117.56,
            'Interest_Expense': 0,
            'Unusual_Expense': 0,
            'Pretax_Income': 2250.236,
            'Income_Tax': 700.046,
            'Net_Income': 1550.19,
            'Pref_Dividends': 0,
            'Net_Income_Common': 1550.19,
            'Basic_Shares': 125.5,
            'Diluted_Shares': 127.1,
        },
        '2025': {
            'Revenue': 10588.126,
            'COGS': 4317.315,
            'Gross_Profit': 6270.811,
            'SGA': 3762.379,
            'Other_OpEx': 2.735,
            'EBITDA': 2952.221,
            'DA': 446.524,
            'EBIT': 2505.697,
            'NonOp_Income': 70.38,
            'Interest_Expense': 0,
            'Unusual_Expense': 0,
            'Pretax_Income': 2576.077,
            'Income_Tax': 761.461,
            'Net_Income': 1814.616,
            'Pref_Dividends': 0,
            'Net_Income_Common': 1814.616,
            'Basic_Shares': 122.0,
            'Diluted_Shares': 124.0,
        },
    }
    
    # Bloomberg Consensus Estimates
    estimates = {
        '2026E': {
            'Revenue': 11040.29,
            'COGS': 4797.78,
            'Gross_Profit': 6242.51,
            'SGA': 3919.31,
            'Other_OpEx': 50.0,
            'EBITDA': 2693.14,
            'DA': 501.42,
            'EBIT': 2191.72,
            'NonOp_Income': 70.0,
            'Interest_Expense': 0,
            'Unusual_Expense': 0,
            'Pretax_Income': 2261.72,
            'Income_Tax': 707.56,
            'Net_Income': 1554.16,
            'Pref_Dividends': 0,
            'Net_Income_Common': 1554.16,
            'Basic_Shares': 119.0,
            'Diluted_Shares': 121.0,
        },
        '2027E': {
            'Revenue': 11537.65,
            'COGS': 5176.02,
            'Gross_Profit': 6361.63,
            'SGA': 4095.87,
            'Other_OpEx': 50.0,
            'EBITDA': 2603.90,
            'DA': 548.97,
            'EBIT': 2054.93,
            'NonOp_Income': 70.0,
            'Interest_Expense': 0,
            'Unusual_Expense': 0,
            'Pretax_Income': 2124.93,
            'Income_Tax': 661.61,
            'Net_Income': 1463.32,
            'Pref_Dividends': 0,
            'Net_Income_Common': 1463.32,
            'Basic_Shares': 116.0,
            'Diluted_Shares': 118.0,
        },
    }
    
    # Balance Sheet Data
    balance_data = {
        '2020': {
            'Cash': 1093.505,
            'Receivables': 40.219,
            'Inventory': 518.513,
            'Other_CA': 155.701,
            'Total_CA': 1807.938,
            'Gross_PPE': 1361.357,
            'Accum_Depr': 0,
            'Net_PPE': 1361.357,
            'Investments': 0,
            'LT_Receivable': 0,
            'Intangibles': 0,
            'Deferred_Tax_A': 31.435,
            'Other_Assets': 80.624,
            'Total_NonCA': 1473.416,
            'Total_Assets': 3281.354,
            'ST_Debt': 128.497,
            'AP': 79.997,
            'Tax_Payable': 26.436,
            'Other_CL': 385.488,
            'Total_CL': 620.418,
            'LT_Debt': 611.464,
            'Provisions': 0,
            'Deferred_Tax_L': 0,
            'Other_LTL': 97.254,
            'Total_NonCL': 708.718,
            'Total_Liab': 1329.136,
            'Pref_Stock': 0,
            'Stockholder_Equity': 1952.218,
            'Total_SH_Equity': 1952.218,
            'Minority_Int': 0,
            'Total_Equity': 1952.218,
        },
        '2021': {
            'Cash': 1150.517,
            'Receivables': 62.399,
            'Inventory': 647.23,
            'Other_CA': 264.233,
            'Total_CA': 2124.379,
            'Gross_PPE': 1480.522,
            'Accum_Depr': 0,
            'Net_PPE': 1480.522,
            'Investments': 0,
            'LT_Receivable': 0,
            'Intangibles': 0,
            'Deferred_Tax_A': 6.731,
            'Other_Assets': 573.583,
            'Total_NonCA': 2060.836,
            'Total_Assets': 4185.215,
            'ST_Debt': 166.091,
            'AP': 172.246,
            'Tax_Payable': 8.357,
            'Other_CL': 536.484,
            'Total_CL': 883.178,
            'LT_Debt': 632.59,
            'Provisions': 0,
            'Deferred_Tax_L': 0,
            'Other_LTL': 110.881,
            'Total_NonCL': 743.471,
            'Total_Liab': 1626.649,
            'Pref_Stock': 0,
            'Stockholder_Equity': 2558.566,
            'Total_SH_Equity': 2558.566,
            'Minority_Int': 0,
            'Total_Equity': 2558.566,
        },
        '2022': {
            'Cash': 1259.871,
            'Receivables': 77.001,
            'Inventory': 966.481,
            'Other_CA': 311.5,
            'Total_CA': 2614.853,
            'Gross_PPE': 1731.253,
            'Accum_Depr': 0,
            'Net_PPE': 1731.253,
            'Investments': 0,
            'LT_Receivable': 0,
            'Intangibles': 0,
            'Deferred_Tax_A': 6.091,
            'Other_Assets': 590.281,
            'Total_NonCA': 2327.625,
            'Total_Assets': 4942.478,
            'ST_Debt': 188.996,
            'AP': 289.728,
            'Tax_Payable': 133.852,
            'Other_CL': 792.758,
            'Total_CL': 1405.334,
            'LT_Debt': 692.056,
            'Provisions': 0,
            'Deferred_Tax_L': 0,
            'Other_LTL': 105.042,
            'Total_NonCL': 797.098,
            'Total_Liab': 2202.432,
            'Pref_Stock': 0,
            'Stockholder_Equity': 2740.046,
            'Total_SH_Equity': 2740.046,
            'Minority_Int': 0,
            'Total_Equity': 2740.046,
        },
        '2023': {
            'Cash': 1154.867,
            'Receivables': 132.906,
            'Inventory': 1447.367,
            'Other_CA': 424.313,
            'Total_CA': 3159.453,
            'Gross_PPE': 2239.033,
            'Accum_Depr': 0,
            'Net_PPE': 2239.033,
            'Investments': 0,
            'LT_Receivable': 0,
            'Intangibles': 0,
            'Deferred_Tax_A': 6.402,
            'Other_Assets': 202.15,
            'Total_NonCA': 2447.585,
            'Total_Assets': 5607.038,
            'ST_Debt': 207.972,
            'AP': 172.732,
            'Tax_Payable': 174.221,
            'Other_CL': 937.273,
            'Total_CL': 1492.198,
            'LT_Debt': 862.362,
            'Provisions': 0,
            'Deferred_Tax_L': 0,
            'Other_LTL': 103.679,
            'Total_NonCL': 966.041,
            'Total_Liab': 2458.239,
            'Pref_Stock': 0,
            'Stockholder_Equity': 3148.799,
            'Total_SH_Equity': 3148.799,
            'Minority_Int': 0,
            'Total_Equity': 3148.799,
        },
        '2024': {
            'Cash': 2243.971,
            'Receivables': 124.769,
            'Inventory': 1323.602,
            'Other_CA': 368.235,
            'Total_CA': 4060.577,
            'Gross_PPE': 2811.421,
            'Accum_Depr': 0,
            'Net_PPE': 2811.421,
            'Investments': 0,
            'LT_Receivable': 0,
            'Intangibles': 0,
            'Deferred_Tax_A': 9.176,
            'Other_Assets': 210.767,
            'Total_NonCA': 3031.364,
            'Total_Assets': 7091.941,
            'ST_Debt': 249.27,
            'AP': 348.441,
            'Tax_Payable': 12.098,
            'Other_CL': 1021.452,
            'Total_CL': 1631.261,
            'LT_Debt': 1154.012,
            'Provisions': 0,
            'Deferred_Tax_L': 0,
            'Other_LTL': 74.587,
            'Total_NonCL': 1228.599,
            'Total_Liab': 2859.86,
            'Pref_Stock': 0,
            'Stockholder_Equity': 4232.081,
            'Total_SH_Equity': 4232.081,
            'Minority_Int': 0,
            'Total_Equity': 4232.081,
        },
        '2025': {
            'Cash': 1984.336,
            'Receivables': 120.173,
            'Inventory': 1442.081,
            'Other_CA': 433.712,
            'Total_CA': 3980.302,
            'Gross_PPE': 3196.873,
            'Accum_Depr': 0,
            'Net_PPE': 3196.873,
            'Investments': 0,
            'LT_Receivable': 0,
            'Intangibles': 0,
            'Deferred_Tax_A': 17.085,
            'Other_Assets': 409.032,
            'Total_NonCA': 3622.99,
            'Total_Assets': 7603.292,
            'ST_Debt': 275.154,
            'AP': 271.406,
            'Tax_Payable': 183.126,
            'Other_CL': 1109.944,
            'Total_CL': 1839.63,
            'LT_Debt': 1300.637,
            'Provisions': 0,
            'Deferred_Tax_L': 0,
            'Other_LTL': 138.978,
            'Total_NonCL': 1439.615,
            'Total_Liab': 3279.245,
            'Pref_Stock': 0,
            'Stockholder_Equity': 4324.047,
            'Total_SH_Equity': 4324.047,
            'Minority_Int': 0,
            'Total_Equity': 4324.047,
        },
    }
    
    # Pro forma Balance Sheet projections
    balance_proforma = {
        '2026E': {
            'Cash': 2100.0,
            'Receivables': 125.0,
            'Inventory': 1500.0,
            'Other_CA': 450.0,
            'Total_CA': 4175.0,
            'Gross_PPE': 3700.0,
            'Accum_Depr': 0,
            'Net_PPE': 3700.0,
            'Investments': 0,
            'LT_Receivable': 0,
            'Intangibles': 0,
            'Deferred_Tax_A': 18.0,
            'Other_Assets': 420.0,
            'Total_NonCA': 4138.0,
            'Total_Assets': 8313.0,
            'ST_Debt': 290.0,
            'AP': 290.0,
            'Tax_Payable': 170.0,
            'Other_CL': 1150.0,
            'Total_CL': 1900.0,
            'LT_Debt': 1350.0,
            'Provisions': 0,
            'Deferred_Tax_L': 0,
            'Other_LTL': 145.0,
            'Total_NonCL': 1495.0,
            'Total_Liab': 3395.0,
            'Pref_Stock': 0,
            'Stockholder_Equity': 4918.0,
            'Total_SH_Equity': 4918.0,
            'Minority_Int': 0,
            'Total_Equity': 4918.0,
        },
        '2027E': {
            'Cash': 2200.0,
            'Receivables': 130.0,
            'Inventory': 1550.0,
            'Other_CA': 470.0,
            'Total_CA': 4350.0,
            'Gross_PPE': 4200.0,
            'Accum_Depr': 0,
            'Net_PPE': 4200.0,
            'Investments': 0,
            'LT_Receivable': 0,
            'Intangibles': 0,
            'Deferred_Tax_A': 19.0,
            'Other_Assets': 430.0,
            'Total_NonCA': 4649.0,
            'Total_Assets': 8999.0,
            'ST_Debt': 305.0,
            'AP': 305.0,
            'Tax_Payable': 160.0,
            'Other_CL': 1180.0,
            'Total_CL': 1950.0,
            'LT_Debt': 1400.0,
            'Provisions': 0,
            'Deferred_Tax_L': 0,
            'Other_LTL': 150.0,
            'Total_NonCL': 1550.0,
            'Total_Liab': 3500.0,
            'Pref_Stock': 0,
            'Stockholder_Equity': 5499.0,
            'Total_SH_Equity': 5499.0,
            'Minority_Int': 0,
            'Total_Equity': 5499.0,
        },
    }
    
    # Cash Flow Data
    cashflow_data = {
        '2021': {
            'Net_Income': 588.913,
            'DA': 185.478,
            'Deferred_Tax': 0,
            'Other_Funds': 51.489,
            'FFO': 825.88,
            'Chg_Receivables': -22.18,
            'Chg_Inventory': -128.717,
            'Chg_AP': 92.249,
            'Other_WC': 36.106,
            'Chg_WC': -22.544,
            'CFO': 803.336,
            'CapEx': -229.226,
            'Acquisitions': 0,
            'Sale_Assets': 0,
            'Investments': 0,
            'Other_Invest': -466.306,
            'CFI': -695.532,
            'Chg_ST_Debt': 0,
            'Chg_LT_Debt': 0,
            'Debt_Net': 0,
            'Pref_Div': 0,
            'Common_Div': 0,
            'Cash_Div': 0,
            'Stock_Change': -80.788,
            'Other_Fin': 29.996,
            'CFF': -50.792,
            'Net_Change': 57.012,
        },
        '2022': {
            'Net_Income': 975.322,
            'DA': 224.206,
            'Deferred_Tax': 0,
            'Other_Funds': 86.46,
            'FFO': 1285.988,
            'Chg_Receivables': -14.602,
            'Chg_Inventory': -319.251,
            'Chg_AP': 117.482,
            'Other_WC': 319.491,
            'Chg_WC': 103.12,
            'CFO': 1389.108,
            'CapEx': -394.502,
            'Acquisitions': 0,
            'Sale_Assets': 0,
            'Investments': 0,
            'Other_Invest': -33.389,
            'CFI': -427.891,
            'Chg_ST_Debt': 0,
            'Chg_LT_Debt': 0,
            'Debt_Net': 0,
            'Pref_Div': 0,
            'Common_Div': 0,
            'Cash_Div': 0,
            'Stock_Change': -844.217,
            'Other_Fin': -7.646,
            'CFF': -851.863,
            'Net_Change': 109.354,
        },
        '2023': {
            'Net_Income': 854.8,
            'DA': 291.791,
            'Deferred_Tax': 0,
            'Other_Funds': 401.326,
            'FFO': 1547.917,
            'Chg_Receivables': -55.905,
            'Chg_Inventory': -480.886,
            'Chg_AP': -116.996,
            'Other_WC': 71.333,
            'Chg_WC': -581.454,
            'CFO': 966.463,
            'CapEx': -638.657,
            'Acquisitions': 0,
            'Sale_Assets': 0,
            'Investments': 0,
            'Other_Invest': 68.72,
            'CFI': -569.937,
            'Chg_ST_Debt': 0,
            'Chg_LT_Debt': 0,
            'Debt_Net': 0,
            'Pref_Div': 0,
            'Common_Div': 0,
            'Cash_Div': 0,
            'Stock_Change': -467.455,
            'Other_Fin': -34.075,
            'CFF': -501.53,
            'Net_Change': -105.004,
        },
        '2024': {
            'Net_Income': 1550.19,
            'DA': 379.384,
            'Deferred_Tax': 0,
            'Other_Funds': 163.852,
            'FFO': 2093.426,
            'Chg_Receivables': 8.137,
            'Chg_Inventory': 123.765,
            'Chg_AP': 175.709,
            'Other_WC': -104.873,
            'Chg_WC': 202.738,
            'CFO': 2296.164,
            'CapEx': -651.865,
            'Acquisitions': 0,
            'Sale_Assets': 0,
            'Investments': 0,
            'Other_Invest': -2.267,
            'CFI': -654.132,
            'Chg_ST_Debt': 0,
            'Chg_LT_Debt': 0,
            'Debt_Net': 0,
            'Pref_Div': 0,
            'Common_Div': 0,
            'Cash_Div': 0,
            'Stock_Change': -548.796,
            'Other_Fin': -4.132,
            'CFF': -552.928,
            'Net_Change': 1089.104,
        },
        '2025': {
            'Net_Income': 1814.616,
            'DA': 446.524,
            'Deferred_Tax': 0,
            'Other_Funds': 16.252,
            'FFO': 2277.392,
            'Chg_Receivables': 4.596,
            'Chg_Inventory': -118.479,
            'Chg_AP': -77.035,
            'Other_WC': 186.239,
            'Chg_WC': -4.679,
            'CFO': 2272.713,
            'CapEx': -689.232,
            'Acquisitions': 0,
            'Sale_Assets': 0,
            'Investments': 0,
            'Other_Invest': -108.942,
            'CFI': -798.174,
            'Chg_ST_Debt': 0,
            'Chg_LT_Debt': 0,
            'Debt_Net': 0,
            'Pref_Div': 0,
            'Common_Div': 0,
            'Cash_Div': 0,
            'Stock_Change': -1652.476,
            'Other_Fin': -81.698,
            'CFF': -1734.174,
            'Net_Change': -259.635,
        },
    }
    
    # Pro forma Cash Flow
    cashflow_proforma = {
        '2026E': {
            'Net_Income': 1554.16,
            'DA': 501.42,
            'Deferred_Tax': 0,
            'Other_Funds': 50.0,
            'FFO': 2105.58,
            'Chg_Receivables': -4.83,
            'Chg_Inventory': -57.92,
            'Chg_AP': 18.59,
            'Other_WC': -40.06,
            'Chg_WC': -84.22,
            'CFO': 2021.36,
            'CapEx': -720.0,
            'Acquisitions': 0,
            'Sale_Assets': 0,
            'Investments': 0,
            'Other_Invest': -50.0,
            'CFI': -770.0,
            'Chg_ST_Debt': 0,
            'Chg_LT_Debt': 0,
            'Debt_Net': 0,
            'Pref_Div': 0,
            'Common_Div': 0,
            'Cash_Div': 0,
            'Stock_Change': -800.0,
            'Other_Fin': -50.0,
            'CFF': -850.0,
            'Net_Change': 401.36,
        },
        '2027E': {
            'Net_Income': 1463.32,
            'DA': 548.97,
            'Deferred_Tax': 0,
            'Other_Funds': 50.0,
            'FFO': 2062.29,
            'Chg_Receivables': -5.0,
            'Chg_Inventory': -50.0,
            'Chg_AP': 15.0,
            'Other_WC': -30.0,
            'Chg_WC': -70.0,
            'CFO': 1992.29,
            'CapEx': -750.0,
            'Acquisitions': 0,
            'Sale_Assets': 0,
            'Investments': 0,
            'Other_Invest': -50.0,
            'CFI': -800.0,
            'Chg_ST_Debt': 0,
            'Chg_LT_Debt': 0,
            'Debt_Net': 0,
            'Pref_Div': 0,
            'Common_Div': 0,
            'Cash_Div': 0,
            'Stock_Change': -800.0,
            'Other_Fin': -50.0,
            'CFF': -850.0,
            'Net_Change': 342.29,
        },
    }
    
    # WACC Components
    wacc_data = {
        'Beta': 1.25,
        'Risk_Free_Rate': 0.043,
        'Market_Return': 0.10,
        'Cost_of_Debt': 0.05,
        'Tax_Rate': 0.296,
        'Equity_Pct': 0.73,
        'Debt_Pct': 0.27,
    }
    
    # =========================================================================
    # UPDATE INCOME STATEMENT
    # =========================================================================
    ws = wb['IncomeStatement']
    
    # Update company name
    ws['A1'] = 'Lululemon Athletica Inc.'
    ws['A1'].font = Font(bold=True, size=14)
    
    # Update year headers (row 3 and row 4)
    year_labels_is = ['2021', '2022', '2023', '2024', '2025', '2026E', '2027E']
    for i, year in enumerate(year_labels_is):
        ws.cell(row=3, column=3+i, value=year)
        ws.cell(row=4, column=3+i, value=year)
    
    # Fill historical data
    years_historical = ['2021', '2022', '2023', '2024', '2025']
    for i, year in enumerate(years_historical):
        col = 3 + i  # Column C onwards
        d = income_data[year]
        ws.cell(row=4, column=col, value=d['Revenue'])
        ws.cell(row=5, column=col, value=d['COGS'])
        ws.cell(row=6, column=col, value=d['Gross_Profit'])
        ws.cell(row=7, column=col, value=d['SGA'])
        ws.cell(row=8, column=col, value=d['Other_OpEx'])
        ws.cell(row=9, column=col, value=d['EBITDA'])
        ws.cell(row=10, column=col, value=d['DA'])
        ws.cell(row=11, column=col, value=d['EBIT'])
        ws.cell(row=12, column=col, value=d['NonOp_Income'])
        ws.cell(row=13, column=col, value=d['Interest_Expense'])
        ws.cell(row=14, column=col, value=d['Unusual_Expense'])
        ws.cell(row=15, column=col, value=d['Pretax_Income'])
        ws.cell(row=16, column=col, value=d['Income_Tax'])
        ws.cell(row=17, column=col, value=0)
        ws.cell(row=18, column=col, value=d['Net_Income'])
        ws.cell(row=19, column=col, value=d['Pref_Dividends'])
        ws.cell(row=20, column=col, value=d['Net_Income_Common'])
        ws.cell(row=21, column=col, value=d['Basic_Shares'])
        ws.cell(row=22, column=col, value=d['Diluted_Shares'])
        ws.cell(row=23, column=col, value=d['Net_Income_Common'] / d['Basic_Shares'])
        ws.cell(row=24, column=col, value=d['Net_Income_Common'] / d['Diluted_Shares'])
    
    # Fill pro forma estimates
    for i, year in enumerate(['2026E', '2027E']):
        col = 8 + i  # Column H and I
        d = estimates[year]
        ws.cell(row=4, column=col, value=d['Revenue'])
        ws.cell(row=5, column=col, value=d['COGS'])
        ws.cell(row=6, column=col, value=d['Gross_Profit'])
        ws.cell(row=7, column=col, value=d['SGA'])
        ws.cell(row=8, column=col, value=d['Other_OpEx'])
        ws.cell(row=9, column=col, value=d['EBITDA'])
        ws.cell(row=10, column=col, value=d['DA'])
        ws.cell(row=11, column=col, value=d['EBIT'])
        ws.cell(row=12, column=col, value=d['NonOp_Income'])
        ws.cell(row=13, column=col, value=d['Interest_Expense'])
        ws.cell(row=14, column=col, value=d['Unusual_Expense'])
        ws.cell(row=15, column=col, value=d['Pretax_Income'])
        ws.cell(row=16, column=col, value=d['Income_Tax'])
        ws.cell(row=17, column=col, value=0)
        ws.cell(row=18, column=col, value=d['Net_Income'])
        ws.cell(row=19, column=col, value=d['Pref_Dividends'])
        ws.cell(row=20, column=col, value=d['Net_Income_Common'])
        ws.cell(row=21, column=col, value=d['Basic_Shares'])
        ws.cell(row=22, column=col, value=d['Diluted_Shares'])
        ws.cell(row=23, column=col, value=d['Net_Income_Common'] / d['Basic_Shares'])
        ws.cell(row=24, column=col, value=d['Net_Income_Common'] / d['Diluted_Shares'])
    
    print("✓ Income Statement updated with actual years (2021-2027E)")
    
    # =========================================================================
    # UPDATE BALANCE SHEET
    # =========================================================================
    ws = wb['BalanceSheet']
    ws['A1'] = 'Lululemon Athletica Inc.'
    ws['A1'].font = Font(bold=True, size=14)
    
    # Update year headers (row 3)
    year_labels_bs = ['2020', '2021', '2022', '2023', '2024', '2025', '2026E', '2027E']
    for i, year in enumerate(year_labels_bs):
        ws.cell(row=3, column=2+i, value=year)
    
    # Fill historical data
    bs_years_historical = ['2020', '2021', '2022', '2023', '2024', '2025']
    for i, year in enumerate(bs_years_historical):
        col = 2 + i  # Column B onwards
        d = balance_data[year]
        ws.cell(row=5, column=col, value=d['Cash'])
        ws.cell(row=6, column=col, value=d['Receivables'])
        ws.cell(row=7, column=col, value=d['Inventory'])
        ws.cell(row=8, column=col, value=d['Other_CA'])
        ws.cell(row=9, column=col, value=d['Total_CA'])
        ws.cell(row=10, column=col, value=d['Gross_PPE'])
        ws.cell(row=11, column=col, value=d['Accum_Depr'])
        ws.cell(row=12, column=col, value=d['Net_PPE'])
        ws.cell(row=13, column=col, value=d['Investments'])
        ws.cell(row=14, column=col, value=d['LT_Receivable'])
        ws.cell(row=15, column=col, value=0)
        ws.cell(row=16, column=col, value=d['Deferred_Tax_A'])
        ws.cell(row=17, column=col, value=d['Other_Assets'])
        ws.cell(row=18, column=col, value=d['Total_NonCA'])
        ws.cell(row=19, column=col, value=d['Total_Assets'])
        ws.cell(row=21, column=col, value=d['ST_Debt'])
        ws.cell(row=22, column=col, value=d['AP'])
        ws.cell(row=23, column=col, value=d['Tax_Payable'])
        ws.cell(row=24, column=col, value=d['Other_CL'])
        ws.cell(row=25, column=col, value=d['Total_CL'])
        ws.cell(row=26, column=col, value=d['LT_Debt'])
        ws.cell(row=27, column=col, value=d['Provisions'])
        ws.cell(row=28, column=col, value=d['Deferred_Tax_L'])
        ws.cell(row=29, column=col, value=d['Other_LTL'])
        ws.cell(row=30, column=col, value=d['Total_NonCL'])
        ws.cell(row=31, column=col, value=d['Total_Liab'])
        ws.cell(row=32, column=col, value=d['Pref_Stock'])
        ws.cell(row=33, column=col, value=d['Stockholder_Equity'])
        ws.cell(row=34, column=col, value=d['Total_SH_Equity'])
        ws.cell(row=35, column=col, value=d['Minority_Int'])
        ws.cell(row=36, column=col, value=d['Total_Equity'])
        # Total Liab & Equity = Total Liabilities + Total Equity (FORMULA)
        ws.cell(row=37, column=col, value=f"={chr(64+col)}31+{chr(64+col)}36")
        # Check: Assets - (Liab + Equity) = 0
        ws.cell(row=38, column=col, value=f"={chr(64+col)}19-{chr(64+col)}37")
    
    # Fill pro forma balance sheet
    for i, year in enumerate(['2026E', '2027E']):
        col = 8 + i  # Column H and I
        d = balance_proforma[year]
        ws.cell(row=5, column=col, value=d['Cash'])
        ws.cell(row=6, column=col, value=d['Receivables'])
        ws.cell(row=7, column=col, value=d['Inventory'])
        ws.cell(row=8, column=col, value=d['Other_CA'])
        ws.cell(row=9, column=col, value=d['Total_CA'])
        ws.cell(row=10, column=col, value=d['Gross_PPE'])
        ws.cell(row=11, column=col, value=d['Accum_Depr'])
        ws.cell(row=12, column=col, value=d['Net_PPE'])
        ws.cell(row=13, column=col, value=d['Investments'])
        ws.cell(row=14, column=col, value=d['LT_Receivable'])
        ws.cell(row=15, column=col, value=0)
        ws.cell(row=16, column=col, value=d['Deferred_Tax_A'])
        ws.cell(row=17, column=col, value=d['Other_Assets'])
        ws.cell(row=18, column=col, value=d['Total_NonCA'])
        ws.cell(row=19, column=col, value=d['Total_Assets'])
        ws.cell(row=21, column=col, value=d['ST_Debt'])
        ws.cell(row=22, column=col, value=d['AP'])
        ws.cell(row=23, column=col, value=d['Tax_Payable'])
        ws.cell(row=24, column=col, value=d['Other_CL'])
        ws.cell(row=25, column=col, value=d['Total_CL'])
        ws.cell(row=26, column=col, value=d['LT_Debt'])
        ws.cell(row=27, column=col, value=d['Provisions'])
        ws.cell(row=28, column=col, value=d['Deferred_Tax_L'])
        ws.cell(row=29, column=col, value=d['Other_LTL'])
        ws.cell(row=30, column=col, value=d['Total_NonCL'])
        ws.cell(row=31, column=col, value=d['Total_Liab'])
        ws.cell(row=32, column=col, value=d['Pref_Stock'])
        ws.cell(row=33, column=col, value=d['Stockholder_Equity'])
        ws.cell(row=34, column=col, value=d['Total_SH_Equity'])
        ws.cell(row=35, column=col, value=d['Minority_Int'])
        ws.cell(row=36, column=col, value=d['Total_Equity'])
        # Total Liab & Equity = Total Liabilities + Total Equity (FORMULA)
        ws.cell(row=37, column=col, value=f"={chr(64+col)}31+{chr(64+col)}36")
        # Check: Assets - (Liab + Equity) = 0
        ws.cell(row=38, column=col, value=f"={chr(64+col)}19-{chr(64+col)}37")
    
    print("✓ Balance Sheet updated with actual years (2020-2027E)")
    print("✓ Assets = Liabilities + Equity formula added for all years")
    
    # =========================================================================
    # UPDATE CASH FLOW STATEMENT
    # =========================================================================
    ws = wb['CashFlow']
    ws['A1'] = 'Lululemon Athletica Inc.'
    ws['A1'].font = Font(bold=True, size=14)
    
    # Update year headers
    year_labels_cf = ['2021', '2022', '2023', '2024', '2025', '2026E', '2027E']
    for i, year in enumerate(year_labels_cf):
        ws.cell(row=3, column=3+i, value=year)
    
    # Fill historical data
    cf_years_historical = ['2021', '2022', '2023', '2024', '2025']
    for i, year in enumerate(cf_years_historical):
        col = 3 + i  # Column C onwards
        d = cashflow_data[year]
        ws.cell(row=5, column=col, value=d['Net_Income'])
        ws.cell(row=6, column=col, value=d['DA'])
        ws.cell(row=7, column=col, value=d['Deferred_Tax'])
        ws.cell(row=8, column=col, value=d['Other_Funds'])
        ws.cell(row=9, column=col, value=d['FFO'])
        ws.cell(row=10, column=col, value=d['Chg_Receivables'])
        ws.cell(row=11, column=col, value=d['Chg_Inventory'])
        ws.cell(row=12, column=col, value=d['Chg_AP'])
        ws.cell(row=13, column=col, value=0)
        ws.cell(row=14, column=col, value=d['Chg_WC'])
        ws.cell(row=15, column=col, value=d['CFO'])
        ws.cell(row=17, column=col, value=d['CapEx'])
        ws.cell(row=18, column=col, value=d['Acquisitions'])
        ws.cell(row=19, column=col, value=d['Sale_Assets'])
        ws.cell(row=20, column=col, value=d['Investments'])
        ws.cell(row=21, column=col, value=d['Other_Invest'])
        ws.cell(row=22, column=col, value=d['CFI'])
        ws.cell(row=24, column=col, value=d['Chg_ST_Debt'])
        ws.cell(row=25, column=col, value=d['Chg_LT_Debt'])
        ws.cell(row=26, column=col, value=d['Debt_Net'])
        ws.cell(row=27, column=col, value=d['Pref_Div'])
        ws.cell(row=28, column=col, value=d['Common_Div'])
        ws.cell(row=29, column=col, value=d['Cash_Div'])
        ws.cell(row=30, column=col, value=d['Stock_Change'])
        ws.cell(row=31, column=col, value=d['Other_Fin'])
        ws.cell(row=32, column=col, value=d['CFF'])
        ws.cell(row=33, column=col, value=d['Net_Change'])
    
    # Fill pro forma cash flow
    for i, year in enumerate(['2026E', '2027E']):
        col = 8 + i  # Column H and I
        d = cashflow_proforma[year]
        ws.cell(row=5, column=col, value=d['Net_Income'])
        ws.cell(row=6, column=col, value=d['DA'])
        ws.cell(row=7, column=col, value=d['Deferred_Tax'])
        ws.cell(row=8, column=col, value=d['Other_Funds'])
        ws.cell(row=9, column=col, value=d['FFO'])
        ws.cell(row=10, column=col, value=d['Chg_Receivables'])
        ws.cell(row=11, column=col, value=d['Chg_Inventory'])
        ws.cell(row=12, column=col, value=d['Chg_AP'])
        ws.cell(row=13, column=col, value=0)
        ws.cell(row=14, column=col, value=d['Chg_WC'])
        ws.cell(row=15, column=col, value=d['CFO'])
        ws.cell(row=17, column=col, value=d['CapEx'])
        ws.cell(row=18, column=col, value=d['Acquisitions'])
        ws.cell(row=19, column=col, value=d['Sale_Assets'])
        ws.cell(row=20, column=col, value=d['Investments'])
        ws.cell(row=21, column=col, value=d['Other_Invest'])
        ws.cell(row=22, column=col, value=d['CFI'])
        ws.cell(row=24, column=col, value=d['Chg_ST_Debt'])
        ws.cell(row=25, column=col, value=d['Chg_LT_Debt'])
        ws.cell(row=26, column=col, value=d['Debt_Net'])
        ws.cell(row=27, column=col, value=d['Pref_Div'])
        ws.cell(row=28, column=col, value=d['Common_Div'])
        ws.cell(row=29, column=col, value=d['Cash_Div'])
        ws.cell(row=30, column=col, value=d['Stock_Change'])
        ws.cell(row=31, column=col, value=d['Other_Fin'])
        ws.cell(row=32, column=col, value=d['CFF'])
        ws.cell(row=33, column=col, value=d['Net_Change'])
    
    print("✓ Cash Flow Statement updated with actual years (2021-2027E)")
    
    # =========================================================================
    # UPDATE WACC MODEL
    # =========================================================================
    ws = wb['WACC']
    ws.cell(row=3, column=4, value=wacc_data['Beta'])
    ws.cell(row=4, column=4, value=wacc_data['Risk_Free_Rate'])
    ws.cell(row=5, column=4, value=wacc_data['Market_Return'])
    
    cost_of_equity = wacc_data['Risk_Free_Rate'] + wacc_data['Beta'] * (wacc_data['Market_Return'] - wacc_data['Risk_Free_Rate'])
    ws.cell(row=6, column=4, value=cost_of_equity)
    
    ws.cell(row=3, column=8, value=wacc_data['Cost_of_Debt'])
    ws.cell(row=4, column=8, value=wacc_data['Tax_Rate'])
    
    after_tax_kd = wacc_data['Cost_of_Debt'] * (1 - wacc_data['Tax_Rate'])
    ws.cell(row=5, column=8, value=after_tax_kd)
    
    ws.cell(row=3, column=11, value=wacc_data['Equity_Pct'])
    ws.cell(row=4, column=11, value=wacc_data['Debt_Pct'])
    
    wacc = wacc_data['Equity_Pct'] * cost_of_equity + wacc_data['Debt_Pct'] * after_tax_kd
    ws.cell(row=7, column=4, value=wacc)
    
    print("✓ WACC Model updated")
    
    # =========================================================================
    # UPDATE DDM MODEL
    # =========================================================================
    ws = wb['DDM']
    ws.cell(row=2, column=5, value=0)  # No dividends
    ws.cell(row=3, column=5, value=0.08)
    ws.cell(row=4, column=5, value=cost_of_equity)
    ws.cell(row=5, column=5, value=170.09)
    ws.cell(row=6, column=5, value=0)
    ws.cell(row=7, column=5, value="N/A - LULU does not pay dividends")
    
    print("✓ DDM Model updated (Note: LULU has no dividends)")
    
    # =========================================================================
    # UPDATE DCF MODEL
    # =========================================================================
    ws = wb['DCF']
    
    # Try to write company name where possible
    try:
        ws['A2'] = 'Lululemon Athletica Inc.'
        ws['A2'].font = Font(bold=True, size=12)
    except:
        pass
    
    # Update year headers (row 4) - skip merged cells
    dcf_years = ['2021', '2022', '2023', '2024', '2025', '2026E', '2027E', '2028E', '2029E', '2030E']
    for i, year in enumerate(dcf_years):
        try:
            cell = ws.cell(row=4, column=2+i)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = year
        except:
            pass
    
    # Fill historical revenues (row 5)
    rev_history = [4401.879, 6256.617, 8110.518, 9619.278, 10588.126]
    for i, rev in enumerate(rev_history):
        try:
            cell = ws.cell(row=5, column=2+i)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = rev
        except:
            pass
    
    # Fill projected revenues
    projected_rev = [11040.29, 11537.65, 12190.48, 12800.0, 13400.0]
    for i, rev in enumerate(projected_rev):
        try:
            cell = ws.cell(row=5, column=7+i)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = rev
        except:
            pass
    
    # Fill COGS
    cogs_history = [1937.888, 2648.052, 3618.178, 4009.873, 4317.315]
    for i, cogs in enumerate(cogs_history):
        try:
            cell = ws.cell(row=6, column=2+i)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = cogs
        except:
            pass
    
    # Fill SG&A
    sga_history = [1609.003, 2225.034, 2757.447, 3397.218, 3762.379]
    for i, sga in enumerate(sga_history):
        try:
            cell = ws.cell(row=7, column=2+i)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = sga
        except:
            pass
    
    # Fill D&A
    da_history = [185.478, 224.206, 291.791, 379.384, 446.524]
    for i, da in enumerate(da_history):
        try:
            cell = ws.cell(row=9, column=2+i)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = da
        except:
            pass
    
    # Fill EBIT
    ebit_history = [819.986, 1333.355, 1328.408, 2132.676, 2505.697]
    for i, ebit in enumerate(ebit_history):
        try:
            cell = ws.cell(row=10, column=2+i)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = ebit
        except:
            pass
    
    # Interest expense - minimal for LULU
    for col in range(2, 12):
        try:
            cell = ws.cell(row=11, column=col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = 0
        except:
            pass
    
    print("✓ DCF Model updated with actual years (2021-2030E)")
    
    # =========================================================================
    # SAVE THE WORKBOOK
    # =========================================================================
    output_path = r"c:\Users\nduta\OneDrive\Desktop\Projects\lulu-lemon-project\Module7_LULU.xlsm"
    wb.save(output_path)
    
    print("\n" + "=" * 70)
    print("✓ Module7.xlsm FULLY UPDATED with Lululemon data!")
    print("=" * 70)
    print(f"\nSaved to: {output_path}")
    print("\nKey Changes Made:")
    print("  • Year headers changed to actual years (2020, 2021, ..., 2027E)")
    print("  • Assets = Liabilities + Equity formula added to all years")
    print("  • All sheets updated with Bloomberg data")
    print("\nData Source: Bloomberg Macro XIDF (1).xlsm")
    
    return wb

if __name__ == "__main__":
    update_module7()
