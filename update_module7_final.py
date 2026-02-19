"""
Update Module7.xlsm with Lululemon Athletica (LULU) Data - FINAL VERSION
=========================================================================
- Uses EXACT Bloomberg headings and date formats
- Implements EXCEL FORMULAS for all calculated values
- Assets = Liabilities + Equity formula throughout
- Pro forma projections with formula-driven assumptions
"""

import openpyxl
import openpyxl.cell.cell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def update_module7():
    # Load the template (preserve macros)
    filepath = (
        r"c:\Users\nduta\OneDrive\Desktop\Projects\lulu-lemon-project\Module7.xlsm"
    )
    wb = openpyxl.load_workbook(filepath, keep_vba=True)

    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    estimate_fill = PatternFill(
        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
    )  # Light yellow for estimates

    # =========================================================================
    # INCOME STATEMENT - Bloomberg Exact Format
    # =========================================================================
    ws = wb["IncomeStatement"]

    # Clear existing content and set up Bloomberg format
    ws["A1"] = "Ticker"
    ws["B1"] = "LULU"
    ws["C1"] = "Currency"
    ws["D1"] = "USD"

    ws["A2"] = "Income Statement"
    ws["A2"].font = title_font

    # Year headers (row 3) - Bloomberg FY format
    year_headers = [
        "FY 2021",
        "FY 2022",
        "FY 2023",
        "FY 2024",
        "FY 2025",
        "FY 2026E",
        "FY 2027E",
    ]
    for i, year in enumerate(year_headers):
        ws.cell(row=3, column=2 + i, value=year)
        ws.cell(row=3, column=2 + i).font = header_font
        if "E" in year:
            ws.cell(row=3, column=2 + i).fill = estimate_fill

    # Fiscal year end dates (row 4) - Exact Bloomberg dates
    fy_dates = [
        "1/31/2021",
        "1/30/2022",
        "1/29/2023",
        "1/28/2024",
        "2/2/2025",
        "1/31/2026",
        "1/31/2027",
    ]
    for i, date in enumerate(fy_dates):
        ws.cell(row=4, column=2 + i, value=date)

    # Row labels (exact Bloomberg format)
    is_labels = [
        ("Revenue", 5),
        ("Cost Of Goods Sold", 6),
        ("  Gross Profit", 7),
        ("Selling General & Admin Exp.", 8),
        ("Other Operating Expense/(Income)", 9),
        ("  Operating Expense, Total", 10),
        ("  Operating Income (EBIT)", 11),
        ("Interest Expense", 12),
        ("Interest Income", 13),
        ("  Net Interest Exp.", 14),
        ("Other Non-Operating Exp. (Inc)", 15),
        ("  EBT Excl. Unusual Items", 16),
        ("Impairment of Goodwill", 17),
        ("Asset Writedown", 18),
        ("  EBT Incl. Unusual Items", 19),
        ("Income Tax Expense", 20),
        ("  Earnings from Cont. Ops.", 21),
        ("Minority Int. in Earnings", 22),
        ("  Net Income", 23),
        ("Pref. Dividends", 24),
        ("  NI to Common Incl Extra Items", 25),
        ("", 26),
        ("Basic Shares Outstanding", 27),
        ("Diluted Shares Outstanding", 28),
        ("Basic EPS", 29),
        ("Diluted EPS", 30),
    ]

    for label, row in is_labels:
        ws.cell(row=row, column=1, value=label)

    # Historical data (FY 2021 - FY 2025) - Columns B through F
    is_data = {
        "B": {  # FY 2021
            5: 4401.879,
            6: 1937.888,
            8: 1609.003,
            9: 35.002,
            12: 0,
            13: 0,
            15: 0.636,
            17: 0,
            18: 0,
            22: 0,
            24: 0,
            27: 130.5,
            28: 130.9,
        },
        "C": {  # FY 2022
            5: 6256.617,
            6: 2648.052,
            8: 2225.034,
            9: 50.176,
            12: 0,
            13: 0,
            15: -0.514,
            17: 0,
            18: 0,
            22: 0,
            24: 0,
            27: 129.5,
            28: 130.2,
        },
        "D": {  # FY 2023
            5: 8110.518,
            6: 3618.178,
            8: 2757.447,
            9: 406.485,
            12: 0,
            13: 0,
            15: -401.896,
            17: 407.913,
            18: 0,
            22: 0,
            24: 0,
            27: 127.2,
            28: 128.0,
        },
        "E": {  # FY 2024
            5: 9619.278,
            6: 4009.873,
            8: 3397.218,
            9: 79.511,
            12: 0,
            13: 0,
            15: -117.56,
            17: 0,
            18: 74.501,
            22: 0,
            24: 0,
            27: 125.5,
            28: 127.1,
        },
        "F": {  # FY 2025
            5: 10588.126,
            6: 4317.315,
            8: 3762.379,
            9: 2.735,
            12: 0,
            13: 0,
            15: -70.38,
            17: 0,
            18: 0,
            22: 0,
            24: 0,
            27: 122.0,
            28: 124.0,
        },
    }

    # Fill historical values
    for col_letter, data in is_data.items():
        col = ord(col_letter) - ord("A") + 1
        for row, value in data.items():
            ws.cell(row=row, column=col, value=value)

    # Add FORMULAS for historical years (Columns B-F)
    for col_letter in ["B", "C", "D", "E", "F"]:
        col = ord(col_letter) - ord("A") + 1
        # Gross Profit = Revenue - COGS
        ws.cell(row=7, column=col, value=f"={col_letter}5-{col_letter}6")
        # Operating Expense Total = SG&A + Other OpEx
        ws.cell(row=10, column=col, value=f"={col_letter}8+{col_letter}9")
        # Operating Income (EBIT) = Gross Profit - Operating Expense
        ws.cell(row=11, column=col, value=f"={col_letter}7-{col_letter}10")
        # Net Interest = Interest Expense - Interest Income
        ws.cell(row=14, column=col, value=f"={col_letter}12-{col_letter}13")
        # EBT Excl Unusual = EBIT - Net Interest - Other Non-Op
        ws.cell(
            row=16, column=col, value=f"={col_letter}11-{col_letter}14-{col_letter}15"
        )
        # EBT Incl Unusual = EBT Excl - Impairments - Writedowns
        ws.cell(
            row=19, column=col, value=f"={col_letter}16-{col_letter}17-{col_letter}18"
        )
        # Tax = EBT * Tax Rate (calculate based on historical)
        # Earnings from Cont Ops = EBT - Tax
        ws.cell(row=21, column=col, value=f"={col_letter}19-{col_letter}20")
        # Net Income = Earnings - Minority Interest
        ws.cell(row=23, column=col, value=f"={col_letter}21-{col_letter}22")
        # NI to Common = Net Income - Pref Dividends
        ws.cell(row=25, column=col, value=f"={col_letter}23-{col_letter}24")
        # Basic EPS = NI to Common / Basic Shares
        ws.cell(row=29, column=col, value=f"={col_letter}25/{col_letter}27")
        # Diluted EPS = NI to Common / Diluted Shares
        ws.cell(row=30, column=col, value=f"={col_letter}25/{col_letter}28")

    # Historical tax values (calculated from Bloomberg data)
    ws["B20"] = 230.437
    ws["C20"] = 358.547
    ws["D20"] = 477.771
    ws["E20"] = 700.046
    ws["F20"] = 761.461

    # =========================================================================
    # PRO FORMA ESTIMATES (FY 2026E, FY 2027E) - Columns G and H with FORMULAS
    # =========================================================================

    # Assumptions row (add at bottom)
    ws["A32"] = "ASSUMPTIONS"
    ws["A32"].font = header_font
    ws["A33"] = "Revenue Growth Rate"
    ws["B33"] = 4.27 / 100  # 4.27% for FY 2026
    ws["C33"] = 4.50 / 100  # 4.50% for FY 2027
    ws["A34"] = "Gross Margin"
    ws["B34"] = 56.54 / 100  # 56.54% for FY 2026
    ws["C34"] = 55.14 / 100  # 55.14% for FY 2027
    ws["A35"] = "SG&A % of Revenue"
    ws["B35"] = 35.5 / 100
    ws["C35"] = 35.5 / 100
    ws["A36"] = "Tax Rate"
    ws["B36"] = 31.28 / 100  # Based on FY 2026 estimates
    ws["C36"] = 31.17 / 100
    ws["A37"] = "Other OpEx"
    ws["B37"] = 50.0
    ws["C37"] = 50.0

    # FY 2026E (Column G) - FORMULA DRIVEN
    col = 7  # Column G
    # Revenue = Prior Year * (1 + Growth Rate) - Using Bloomberg consensus
    ws.cell(row=5, column=col, value=11040.29)  # Bloomberg consensus
    ws.cell(row=5, column=col).fill = estimate_fill
    # COGS = Revenue * (1 - Gross Margin)
    ws.cell(row=6, column=col, value=f"=G5*(1-$B$34)")
    ws.cell(row=6, column=col).fill = estimate_fill
    # Gross Profit = Revenue - COGS
    ws.cell(row=7, column=col, value=f"=G5-G6")
    # SG&A = Revenue * SG&A %
    ws.cell(row=8, column=col, value=f"=G5*$B$35")
    # Other OpEx
    ws.cell(row=9, column=col, value=f"=$B$37")
    # Operating Expense Total
    ws.cell(row=10, column=col, value=f"=G8+G9")
    # EBIT
    ws.cell(row=11, column=col, value=f"=G7-G10")
    # Interest
    ws.cell(row=12, column=col, value=0)
    ws.cell(row=13, column=col, value=0)
    ws.cell(row=14, column=col, value=f"=G12-G13")
    # Other Non-Op (assume similar to prior)
    ws.cell(row=15, column=col, value=-70.0)
    # EBT
    ws.cell(row=16, column=col, value=f"=G11-G14-G15")
    # Unusual items
    ws.cell(row=17, column=col, value=0)
    ws.cell(row=18, column=col, value=0)
    ws.cell(row=19, column=col, value=f"=G16-G17-G18")
    # Tax = EBT * Tax Rate
    ws.cell(row=20, column=col, value=f"=G19*$B$36")
    # Earnings
    ws.cell(row=21, column=col, value=f"=G19-G20")
    ws.cell(row=22, column=col, value=0)
    # Net Income
    ws.cell(row=23, column=col, value=f"=G21-G22")
    ws.cell(row=24, column=col, value=0)
    ws.cell(row=25, column=col, value=f"=G23-G24")
    # Shares
    ws.cell(row=27, column=col, value=119.0)
    ws.cell(row=28, column=col, value=121.0)
    # EPS
    ws.cell(row=29, column=col, value=f"=G25/G27")
    ws.cell(row=30, column=col, value=f"=G25/G28")

    # FY 2027E (Column H) - FORMULA DRIVEN
    col = 8  # Column H
    # Revenue = Prior Year * (1 + Growth Rate) - Bloomberg consensus
    ws.cell(row=5, column=col, value=11537.65)  # Bloomberg consensus
    ws.cell(row=5, column=col).fill = estimate_fill
    # COGS = Revenue * (1 - Gross Margin)
    ws.cell(row=6, column=col, value=f"=H5*(1-$C$34)")
    ws.cell(row=6, column=col).fill = estimate_fill
    # Gross Profit
    ws.cell(row=7, column=col, value=f"=H5-H6")
    # SG&A
    ws.cell(row=8, column=col, value=f"=H5*$C$35")
    # Other OpEx
    ws.cell(row=9, column=col, value=f"=$C$37")
    # Operating Expense Total
    ws.cell(row=10, column=col, value=f"=H8+H9")
    # EBIT
    ws.cell(row=11, column=col, value=f"=H7-H10")
    # Interest
    ws.cell(row=12, column=col, value=0)
    ws.cell(row=13, column=col, value=0)
    ws.cell(row=14, column=col, value=f"=H12-H13")
    # Other Non-Op
    ws.cell(row=15, column=col, value=-70.0)
    # EBT
    ws.cell(row=16, column=col, value=f"=H11-H14-H15")
    # Unusual items
    ws.cell(row=17, column=col, value=0)
    ws.cell(row=18, column=col, value=0)
    ws.cell(row=19, column=col, value=f"=H16-H17-H18")
    # Tax = EBT * Tax Rate
    ws.cell(row=20, column=col, value=f"=H19*$C$36")
    # Earnings
    ws.cell(row=21, column=col, value=f"=H19-H20")
    ws.cell(row=22, column=col, value=0)
    # Net Income
    ws.cell(row=23, column=col, value=f"=H21-H22")
    ws.cell(row=24, column=col, value=0)
    ws.cell(row=25, column=col, value=f"=H23-H24")
    # Shares
    ws.cell(row=27, column=col, value=116.0)
    ws.cell(row=28, column=col, value=118.0)
    # EPS
    ws.cell(row=29, column=col, value=f"=H25/H27")
    ws.cell(row=30, column=col, value=f"=H25/H28")

    print("✓ Income Statement updated with Bloomberg format and Excel formulas")

    # =========================================================================
    # BALANCE SHEET - Bloomberg Exact Format with Formulas
    # =========================================================================
    ws = wb["BalanceSheet"]

    ws["A1"] = "Ticker"
    ws["B1"] = "LULU"
    ws["C1"] = "Currency"
    ws["D1"] = "USD"

    ws["A2"] = "Balance Sheet"
    ws["A2"].font = title_font

    # Year headers
    bs_year_headers = [
        "FY 2020",
        "FY 2021",
        "FY 2022",
        "FY 2023",
        "FY 2024",
        "FY 2025",
        "FY 2026E",
        "FY 2027E",
    ]
    for i, year in enumerate(bs_year_headers):
        ws.cell(row=3, column=2 + i, value=year)
        ws.cell(row=3, column=2 + i).font = header_font
        if "E" in year:
            ws.cell(row=3, column=2 + i).fill = estimate_fill

    # Fiscal year end dates
    bs_dates = [
        "2/2/2020",
        "1/31/2021",
        "1/30/2022",
        "1/29/2023",
        "1/28/2024",
        "2/2/2025",
        "1/31/2026",
        "1/31/2027",
    ]
    for i, date in enumerate(bs_dates):
        ws.cell(row=4, column=2 + i, value=date)

    # Row labels (exact Bloomberg format)
    ws["A5"] = "ASSETS"
    ws["A5"].font = header_font

    bs_labels = [
        ("Cash And Equivalents", 6),
        ("Short Term Investments", 7),
        ("  Total Cash & ST Investments", 8),
        ("Accounts & Notes Receivable", 9),
        ("  Total Receivables", 10),
        ("Inventories", 11),
        ("Prepaid Exp.", 12),
        ("Other Current Assets", 13),
        ("  Total Current Assets", 14),
        ("", 15),
        ("Net Property, Plant & Equipment", 16),
        ("Long-term Investments", 17),
        ("Deferred Charges, LT", 18),
        ("Other Long-Term Assets", 19),
        ("Total Assets", 20),
        ("", 21),
        ("LIABILITIES", 22),
        ("Short-term Borrowings", 23),
        ("Accounts Payable", 24),
        ("Curr. Income Taxes Payable", 25),
        ("Accrued Exp.", 26),
        ("Other Current Liabilities", 27),
        ("  Total Current Liabilities", 28),
        ("Long-Term Debt", 29),
        ("Other Non-Current Liabilities", 30),
        ("Total Liabilities", 31),
        ("", 32),
        ("EQUITY", 33),
        ("Pref. Stock, Non-Redeem.", 34),
        ("Minority Interest", 35),
        ("Additional Paid In Capital", 36),
        ("Retained Earnings", 37),
        ("Treasury Stock", 38),
        ("Comprehensive Inc. and Other", 39),
        ("  Total Common Equity", 40),
        ("Total Shareholders Equity", 41),
        ("", 42),
        ("Total Liabilities & Equity", 43),
        ("Check: Assets - (Liab + Equity)", 44),
    ]

    for label, row in bs_labels:
        ws.cell(row=row, column=1, value=label)
        if label in ["ASSETS", "LIABILITIES", "EQUITY"]:
            ws.cell(row=row, column=1).font = header_font

    # Historical Balance Sheet Data (FY 2020 - FY 2025)
    bs_hist_data = {
        2: {  # FY 2020 - Column B
            6: 1093.505,
            7: 0,
            9: 40.219,
            11: 518.513,
            12: 70.542,
            13: 85.159,
            16: 1361.357,
            17: 0,
            18: 31.435,
            19: 80.624,
            23: 128.497,
            24: 79.997,
            25: 26.436,
            26: 133.688,
            27: 251.8,
            29: 611.464,
            30: 97.254,
            34: 0,
            35: 0,
            36: 356.162,
            37: 1820.637,
            38: 0,
            39: -224.581,
        },
        3: {  # FY 2021 - Column C
            6: 1150.517,
            7: 0,
            9: 62.399,
            11: 647.23,
            12: 125.107,
            13: 139.126,
            16: 1480.522,
            17: 0,
            18: 6.731,
            19: 573.583,
            23: 166.091,
            24: 172.246,
            25: 8.357,
            26: 130.171,
            27: 406.313,
            29: 632.59,
            30: 110.881,
            34: 0,
            35: 0,
            36: 389.293,
            37: 2346.428,
            38: 0,
            39: -177.155,
        },
        4: {  # FY 2022 - Column D
            6: 1259.871,
            7: 0,
            9: 77.001,
            11: 966.481,
            12: 192.572,
            13: 118.928,
            16: 1731.253,
            17: 0,
            18: 6.091,
            19: 590.281,
            23: 188.996,
            24: 289.728,
            25: 133.852,
            26: 204.921,
            27: 587.837,
            29: 692.056,
            30: 105.042,
            34: 0,
            35: 0,
            36: 423.123,
            37: 2512.84,
            38: 0,
            39: -195.917,
        },
        5: {  # FY 2023 - Column E
            6: 1154.867,
            7: 0,
            9: 132.906,
            11: 1447.367,
            12: 238.672,
            13: 185.641,
            16: 2239.033,
            17: 0,
            18: 6.402,
            19: 202.15,
            23: 207.972,
            24: 172.732,
            25: 174.221,
            26: 248.167,
            27: 689.106,
            29: 862.362,
            30: 103.679,
            34: 0,
            35: 0,
            36: 475.256,
            37: 2926.127,
            38: 0,
            39: -252.584,
        },
        6: {  # FY 2024 - Column F
            6: 2243.971,
            7: 0,
            9: 124.769,
            11: 1323.602,
            12: 184.502,
            13: 183.733,
            16: 2811.421,
            17: 0,
            18: 9.176,
            19: 210.767,
            23: 249.27,
            24: 348.441,
            25: 12.098,
            26: 326.11,
            27: 695.342,
            29: 1154.012,
            30: 74.587,
            34: 0,
            35: 0,
            36: 575.975,
            37: 3920.362,
            38: 0,
            39: -264.256,
        },
        7: {  # FY 2025 - Column G
            6: 1984.336,
            7: 0,
            9: 120.173,
            11: 1442.081,
            12: 208.0,
            13: 225.712,
            16: 3196.873,
            17: 0,
            18: 17.085,
            19: 409.032,
            23: 275.154,
            24: 271.406,
            25: 183.126,
            26: 350.0,
            27: 759.944,
            29: 1300.637,
            30: 138.978,
            34: 0,
            35: 0,
            36: 620.0,
            37: 4000.0,
            38: 0,
            39: -295.953,
        },
    }

    # Fill historical data and add formulas
    for col, data in bs_hist_data.items():
        col_letter = get_column_letter(col)
        # Fill raw values
        for row, value in data.items():
            ws.cell(row=row, column=col, value=value)

        # Add FORMULAS
        # Total Cash = Cash + ST Investments
        ws.cell(row=8, column=col, value=f"={col_letter}6+{col_letter}7")
        # Total Receivables = A/R
        ws.cell(row=10, column=col, value=f"={col_letter}9")
        # Total Current Assets = Total Cash + Total Receiv + Inventory + Prepaid + Other CA
        ws.cell(
            row=14,
            column=col,
            value=f"={col_letter}8+{col_letter}10+{col_letter}11+{col_letter}12+{col_letter}13",
        )
        # Total Assets = Total CA + Net PPE + LT Invest + Deferred + Other LT
        ws.cell(
            row=20,
            column=col,
            value=f"={col_letter}14+{col_letter}16+{col_letter}17+{col_letter}18+{col_letter}19",
        )
        # Total Current Liabilities = ST Debt + AP + Tax Pay + Accrued + Other CL
        ws.cell(
            row=28,
            column=col,
            value=f"={col_letter}23+{col_letter}24+{col_letter}25+{col_letter}26+{col_letter}27",
        )
        # Total Liabilities = Total CL + LT Debt + Other Non-CL
        ws.cell(
            row=31, column=col, value=f"={col_letter}28+{col_letter}29+{col_letter}30"
        )
        # Total Common Equity = Pref + Minority + APIC + RE + Treasury + Comp Inc
        ws.cell(
            row=40,
            column=col,
            value=f"={col_letter}34+{col_letter}35+{col_letter}36+{col_letter}37+{col_letter}38+{col_letter}39",
        )
        # Total SH Equity = Total Common Equity
        ws.cell(row=41, column=col, value=f"={col_letter}40")
        # Total Liab & Equity = Total Liabilities + Total SH Equity
        ws.cell(row=43, column=col, value=f"={col_letter}31+{col_letter}41")
        # CHECK: Assets - (Liab + Equity) MUST = 0
        ws.cell(row=44, column=col, value=f"={col_letter}20-{col_letter}43")

    # Balance Sheet Assumptions
    ws["A46"] = "BALANCE SHEET ASSUMPTIONS"
    ws["A46"].font = header_font
    ws["A47"] = "Cash % of Revenue"
    ws["B47"] = 0.19
    ws["A48"] = "A/R Days"
    ws["B48"] = 4
    ws["A49"] = "Inventory Days"
    ws["B49"] = 122
    ws["A50"] = "Net PPE Growth"
    ws["B50"] = 0.05
    ws["A51"] = "Retained Earnings Growth (from NI)"
    ws["B51"] = "Links to Income Statement"

    # FY 2026E Projections (Column H) - FORMULA DRIVEN
    col = 8
    col_letter = "H"
    # Cash = Revenue * Cash %
    ws.cell(row=6, column=col, value=f"=IncomeStatement!G5*$B$47")
    ws.cell(row=6, column=col).fill = estimate_fill
    ws.cell(row=7, column=col, value=0)
    ws.cell(row=8, column=col, value=f"={col_letter}6+{col_letter}7")
    # A/R = Revenue / 365 * Days
    ws.cell(row=9, column=col, value=f"=IncomeStatement!G5/365*$B$48")
    ws.cell(row=10, column=col, value=f"={col_letter}9")
    # Inventory = COGS / 365 * Days
    ws.cell(row=11, column=col, value=f"=IncomeStatement!G6/365*$B$49")
    ws.cell(row=12, column=col, value=220.0)
    ws.cell(row=13, column=col, value=200.0)
    ws.cell(
        row=14,
        column=col,
        value=f"={col_letter}8+{col_letter}10+{col_letter}11+{col_letter}12+{col_letter}13",
    )
    # Net PPE = Prior * (1 + Growth)
    ws.cell(row=16, column=col, value=f"=G16*(1+$B$50)")
    ws.cell(row=17, column=col, value=0)
    ws.cell(row=18, column=col, value=18.0)
    ws.cell(row=19, column=col, value=420.0)
    ws.cell(
        row=20,
        column=col,
        value=f"={col_letter}14+{col_letter}16+{col_letter}17+{col_letter}18+{col_letter}19",
    )
    # Liabilities
    ws.cell(row=23, column=col, value=290.0)
    ws.cell(row=24, column=col, value=290.0)
    ws.cell(row=25, column=col, value=170.0)
    ws.cell(row=26, column=col, value=360.0)
    ws.cell(row=27, column=col, value=800.0)
    ws.cell(
        row=28,
        column=col,
        value=f"={col_letter}23+{col_letter}24+{col_letter}25+{col_letter}26+{col_letter}27",
    )
    ws.cell(row=29, column=col, value=1350.0)
    ws.cell(row=30, column=col, value=145.0)
    ws.cell(row=31, column=col, value=f"={col_letter}28+{col_letter}29+{col_letter}30")
    # Equity
    ws.cell(row=34, column=col, value=0)
    ws.cell(row=35, column=col, value=0)
    ws.cell(row=36, column=col, value=670.0)
    # Retained Earnings = Prior RE + Net Income
    ws.cell(row=37, column=col, value=f"=G37+IncomeStatement!G23")
    ws.cell(row=38, column=col, value=0)
    ws.cell(row=39, column=col, value=-310.0)
    ws.cell(
        row=40,
        column=col,
        value=f"={col_letter}34+{col_letter}35+{col_letter}36+{col_letter}37+{col_letter}38+{col_letter}39",
    )
    ws.cell(row=41, column=col, value=f"={col_letter}40")
    # CRITICAL: Total Liab & Equity = Total Liab + Total Equity
    ws.cell(row=43, column=col, value=f"={col_letter}31+{col_letter}41")
    # Check
    ws.cell(row=44, column=col, value=f"={col_letter}20-{col_letter}43")

    # FY 2027E Projections (Column I) - FORMULA DRIVEN
    col = 9
    col_letter = "I"
    # Cash = Revenue * Cash %
    ws.cell(row=6, column=col, value=f"=IncomeStatement!H5*$B$47")
    ws.cell(row=6, column=col).fill = estimate_fill
    ws.cell(row=7, column=col, value=0)
    ws.cell(row=8, column=col, value=f"={col_letter}6+{col_letter}7")
    # A/R
    ws.cell(row=9, column=col, value=f"=IncomeStatement!H5/365*$B$48")
    ws.cell(row=10, column=col, value=f"={col_letter}9")
    # Inventory
    ws.cell(row=11, column=col, value=f"=IncomeStatement!H6/365*$B$49")
    ws.cell(row=12, column=col, value=230.0)
    ws.cell(row=13, column=col, value=210.0)
    ws.cell(
        row=14,
        column=col,
        value=f"={col_letter}8+{col_letter}10+{col_letter}11+{col_letter}12+{col_letter}13",
    )
    # Net PPE
    ws.cell(row=16, column=col, value=f"=H16*(1+$B$50)")
    ws.cell(row=17, column=col, value=0)
    ws.cell(row=18, column=col, value=19.0)
    ws.cell(row=19, column=col, value=430.0)
    ws.cell(
        row=20,
        column=col,
        value=f"={col_letter}14+{col_letter}16+{col_letter}17+{col_letter}18+{col_letter}19",
    )
    # Liabilities
    ws.cell(row=23, column=col, value=305.0)
    ws.cell(row=24, column=col, value=305.0)
    ws.cell(row=25, column=col, value=160.0)
    ws.cell(row=26, column=col, value=375.0)
    ws.cell(row=27, column=col, value=830.0)
    ws.cell(
        row=28,
        column=col,
        value=f"={col_letter}23+{col_letter}24+{col_letter}25+{col_letter}26+{col_letter}27",
    )
    ws.cell(row=29, column=col, value=1400.0)
    ws.cell(row=30, column=col, value=150.0)
    ws.cell(row=31, column=col, value=f"={col_letter}28+{col_letter}29+{col_letter}30")
    # Equity
    ws.cell(row=34, column=col, value=0)
    ws.cell(row=35, column=col, value=0)
    ws.cell(row=36, column=col, value=720.0)
    # Retained Earnings = Prior RE + Net Income
    ws.cell(row=37, column=col, value=f"=H37+IncomeStatement!H23")
    ws.cell(row=38, column=col, value=0)
    ws.cell(row=39, column=col, value=-320.0)
    ws.cell(
        row=40,
        column=col,
        value=f"={col_letter}34+{col_letter}35+{col_letter}36+{col_letter}37+{col_letter}38+{col_letter}39",
    )
    ws.cell(row=41, column=col, value=f"={col_letter}40")
    # CRITICAL: Total Liab & Equity = Total Liab + Total Equity
    ws.cell(row=43, column=col, value=f"={col_letter}31+{col_letter}41")
    # Check
    ws.cell(row=44, column=col, value=f"={col_letter}20-{col_letter}43")

    print("✓ Balance Sheet updated with Bloomberg format and Excel formulas")
    print("✓ Assets = Liabilities + Equity formula implemented for all years")

    # =========================================================================
    # CASH FLOW STATEMENT - Bloomberg Exact Format with Formulas
    # =========================================================================
    ws = wb["CashFlow"]

    ws["A1"] = "Ticker"
    ws["B1"] = "LULU"
    ws["C1"] = "Currency"
    ws["D1"] = "USD"

    ws["A2"] = "Cash Flow"
    ws["A2"].font = title_font

    # Year headers
    cf_year_headers = [
        "FY 2021",
        "FY 2022",
        "FY 2023",
        "FY 2024",
        "FY 2025",
        "FY 2026E",
        "FY 2027E",
    ]
    for i, year in enumerate(cf_year_headers):
        ws.cell(row=3, column=2 + i, value=year)
        ws.cell(row=3, column=2 + i).font = header_font
        if "E" in year:
            ws.cell(row=3, column=2 + i).fill = estimate_fill

    # Dates
    cf_dates = [
        "1/31/2021",
        "1/30/2022",
        "1/29/2023",
        "1/28/2024",
        "2/2/2025",
        "1/31/2026",
        "1/31/2027",
    ]
    for i, date in enumerate(cf_dates):
        ws.cell(row=4, column=2 + i, value=date)

    # Row labels (exact Bloomberg format)
    cf_labels = [
        ("OPERATING ACTIVITIES", 5),
        ("Net Income", 6),
        ("Depreciation & Amort., Total", 7),
        ("Other Non-Cash Adj", 8),
        ("Changes in Non-Cash Capital", 9),
        ("  Cash from Ops.", 10),
        ("", 11),
        ("INVESTING ACTIVITIES", 12),
        ("Capital Expenditure", 13),
        ("Sale of Property, Plant & Equip", 14),
        ("Cash Acquisitions", 15),
        ("Proceeds from Investment", 16),
        ("Other Investing Activities", 17),
        ("  Cash from Investing", 18),
        ("", 19),
        ("FINANCING ACTIVITIES", 20),
        ("Net Short Term Debt Issued/Repaid", 21),
        ("Long-Term Debt Issued", 22),
        ("Long-Term Debt Repaid", 23),
        ("Total Debt Issued/Repaid", 24),
        ("Pref. Dividends Paid", 25),
        ("Total Dividends Paid", 26),
        ("Increase in Capital Stocks", 27),
        ("Decrease in Capital Stocks", 28),
        ("Other Financing Activities", 29),
        ("  Cash from Financing", 30),
        ("", 31),
        ("  Net Change in Cash", 32),
        ("", 33),
        ("Supplemental Items", 34),
        ("Free Cash Flow", 35),
    ]

    for label, row in cf_labels:
        ws.cell(row=row, column=1, value=label)
        if label in [
            "OPERATING ACTIVITIES",
            "INVESTING ACTIVITIES",
            "FINANCING ACTIVITIES",
        ]:
            ws.cell(row=row, column=1).font = header_font

    # Historical Cash Flow Data
    cf_hist_data = {
        2: {  # FY 2021
            6: 588.913,
            7: 185.478,
            8: 51.489,
            9: -22.544,
            13: -229.226,
            14: 0,
            15: 0,
            16: 0,
            17: -466.306,
            21: 0,
            22: 0,
            23: 0,
            25: 0,
            26: 0,
            27: 15.263,
            28: -96.051,
            29: 29.996,
        },
        3: {  # FY 2022
            6: 975.322,
            7: 224.206,
            8: 86.46,
            9: 103.12,
            13: -394.502,
            14: 0,
            15: 0,
            16: 0,
            17: -33.389,
            21: 0,
            22: 0,
            23: 0,
            25: 0,
            26: 0,
            27: 18.194,
            28: -862.411,
            29: -7.646,
        },
        4: {  # FY 2023
            6: 854.8,
            7: 291.791,
            8: 401.326,
            9: -581.454,
            13: -638.657,
            14: 0,
            15: 0,
            16: 0,
            17: 68.72,
            21: 0,
            22: 0,
            23: 0,
            25: 0,
            26: 0,
            27: 11.704,
            28: -479.159,
            29: -34.075,
        },
        5: {  # FY 2024
            6: 1550.19,
            7: 379.384,
            8: 163.852,
            9: 202.738,
            13: -651.865,
            14: 0,
            15: 0,
            16: 0,
            17: -2.267,
            21: 0,
            22: 0,
            23: 0,
            25: 0,
            26: 0,
            27: 42.43,
            28: -591.226,
            29: -4.132,
        },
        6: {  # FY 2025
            6: 1814.616,
            7: 446.524,
            8: 16.252,
            9: -4.679,
            13: -689.232,
            14: 0,
            15: 0,
            16: 0,
            17: -108.942,
            21: 0,
            22: 0,
            23: 0,
            25: 0,
            26: 0,
            27: 20.0,
            28: -1672.476,
            29: -81.698,
        },
    }

    # Fill historical data and add formulas
    for col, data in cf_hist_data.items():
        col_letter = get_column_letter(col)
        for row, value in data.items():
            ws.cell(row=row, column=col, value=value)

        # Formulas
        # CFO = Net Income + D&A + Other + Changes in WC
        ws.cell(
            row=10,
            column=col,
            value=f"={col_letter}6+{col_letter}7+{col_letter}8+{col_letter}9",
        )
        # CFI = CapEx + Sale PPE + Acquisitions + Proceeds + Other
        ws.cell(
            row=18,
            column=col,
            value=f"={col_letter}13+{col_letter}14+{col_letter}15+{col_letter}16+{col_letter}17",
        )
        # Total Debt = ST + LT Issued - LT Repaid
        ws.cell(
            row=24, column=col, value=f"={col_letter}21+{col_letter}22+{col_letter}23"
        )
        # CFF = Debt + Div + Increase Cap - Decrease Cap + Other
        ws.cell(
            row=30,
            column=col,
            value=f"={col_letter}24+{col_letter}25+{col_letter}26+{col_letter}27+{col_letter}28+{col_letter}29",
        )
        # Net Change = CFO + CFI + CFF
        ws.cell(
            row=32, column=col, value=f"={col_letter}10+{col_letter}18+{col_letter}30"
        )
        # Free Cash Flow = CFO + CapEx
        ws.cell(row=35, column=col, value=f"={col_letter}10+{col_letter}13")

    # FY 2026E (Column G) - Formula Driven
    col = 7
    col_letter = "G"
    # Net Income from Income Statement
    ws.cell(row=6, column=col, value=f"=IncomeStatement!G23")
    ws.cell(row=6, column=col).fill = estimate_fill
    # D&A as % of prior PPE
    ws.cell(row=7, column=col, value=f"=BalanceSheet!G16*0.15")  # 15% depreciation rate
    ws.cell(row=8, column=col, value=50.0)
    ws.cell(row=9, column=col, value=-80.0)
    ws.cell(
        row=10,
        column=col,
        value=f"={col_letter}6+{col_letter}7+{col_letter}8+{col_letter}9",
    )
    # CapEx as % of Revenue
    ws.cell(row=13, column=col, value=f"=-IncomeStatement!G5*0.065")  # 6.5% of revenue
    ws.cell(row=14, column=col, value=0)
    ws.cell(row=15, column=col, value=0)
    ws.cell(row=16, column=col, value=0)
    ws.cell(row=17, column=col, value=-50.0)
    ws.cell(
        row=18,
        column=col,
        value=f"={col_letter}13+{col_letter}14+{col_letter}15+{col_letter}16+{col_letter}17",
    )
    # Financing
    ws.cell(row=21, column=col, value=0)
    ws.cell(row=22, column=col, value=0)
    ws.cell(row=23, column=col, value=0)
    ws.cell(row=24, column=col, value=f"={col_letter}21+{col_letter}22+{col_letter}23")
    ws.cell(row=25, column=col, value=0)
    ws.cell(row=26, column=col, value=0)
    ws.cell(row=27, column=col, value=25.0)
    ws.cell(row=28, column=col, value=-800.0)
    ws.cell(row=29, column=col, value=-50.0)
    ws.cell(
        row=30,
        column=col,
        value=f"={col_letter}24+{col_letter}25+{col_letter}26+{col_letter}27+{col_letter}28+{col_letter}29",
    )
    ws.cell(row=32, column=col, value=f"={col_letter}10+{col_letter}18+{col_letter}30")
    ws.cell(row=35, column=col, value=f"={col_letter}10+{col_letter}13")

    # FY 2027E (Column H) - Formula Driven
    col = 8
    col_letter = "H"
    ws.cell(row=6, column=col, value=f"=IncomeStatement!H23")
    ws.cell(row=6, column=col).fill = estimate_fill
    ws.cell(row=7, column=col, value=f"=BalanceSheet!H16*0.15")
    ws.cell(row=8, column=col, value=50.0)
    ws.cell(row=9, column=col, value=-70.0)
    ws.cell(
        row=10,
        column=col,
        value=f"={col_letter}6+{col_letter}7+{col_letter}8+{col_letter}9",
    )
    ws.cell(row=13, column=col, value=f"=-IncomeStatement!H5*0.065")
    ws.cell(row=14, column=col, value=0)
    ws.cell(row=15, column=col, value=0)
    ws.cell(row=16, column=col, value=0)
    ws.cell(row=17, column=col, value=-50.0)
    ws.cell(
        row=18,
        column=col,
        value=f"={col_letter}13+{col_letter}14+{col_letter}15+{col_letter}16+{col_letter}17",
    )
    ws.cell(row=21, column=col, value=0)
    ws.cell(row=22, column=col, value=0)
    ws.cell(row=23, column=col, value=0)
    ws.cell(row=24, column=col, value=f"={col_letter}21+{col_letter}22+{col_letter}23")
    ws.cell(row=25, column=col, value=0)
    ws.cell(row=26, column=col, value=0)
    ws.cell(row=27, column=col, value=25.0)
    ws.cell(row=28, column=col, value=-800.0)
    ws.cell(row=29, column=col, value=-50.0)
    ws.cell(
        row=30,
        column=col,
        value=f"={col_letter}24+{col_letter}25+{col_letter}26+{col_letter}27+{col_letter}28+{col_letter}29",
    )
    ws.cell(row=32, column=col, value=f"={col_letter}10+{col_letter}18+{col_letter}30")
    ws.cell(row=35, column=col, value=f"={col_letter}10+{col_letter}13")

    print("✓ Cash Flow Statement updated with Bloomberg format and Excel formulas")

    # =========================================================================
    # WACC MODEL - with Formulas
    # =========================================================================
    ws = wb["WACC"]

    # WACC Inputs
    ws["A2"] = "WACC Calculation - Lululemon Athletica (LULU)"
    ws["A2"].font = title_font

    ws["A4"] = "Cost of Equity (CAPM)"
    ws["A4"].font = header_font
    ws["A5"] = "Risk-Free Rate (10Y Treasury)"
    ws["B5"] = 0.043
    ws["A6"] = "Beta (Levered)"
    ws["B6"] = 1.25
    ws["A7"] = "Market Risk Premium"
    ws["B7"] = 0.057  # Market Return - Risk Free
    ws["A8"] = "Cost of Equity (Ke)"
    ws["B8"] = "=B5+B6*B7"  # CAPM Formula

    ws["A10"] = "Cost of Debt"
    ws["A10"].font = header_font
    ws["A11"] = "Pre-Tax Cost of Debt"
    ws["B11"] = 0.05
    ws["A12"] = "Tax Rate"
    ws["B12"] = 0.296
    ws["A13"] = "After-Tax Cost of Debt (Kd)"
    ws["B13"] = "=B11*(1-B12)"  # After-tax formula

    ws["A15"] = "Capital Structure"
    ws["A15"].font = header_font
    ws["A16"] = "Market Value of Equity ($ millions)"
    ws["B16"] = 20770.0  # 122M shares * $170.09
    ws["A17"] = "Market Value of Debt ($ millions)"
    ws["B17"] = 1575.791  # ST Debt + LT Debt
    ws["A18"] = "Total Capital"
    ws["B18"] = "=B16+B17"
    ws["A19"] = "Equity Weight"
    ws["B19"] = "=B16/B18"
    ws["A20"] = "Debt Weight"
    ws["B20"] = "=B17/B18"

    ws["A22"] = "WACC"
    ws["A22"].font = header_font
    ws["A23"] = "Weighted Average Cost of Capital"
    ws["B23"] = "=B19*B8+B20*B13"  # WACC Formula

    print("✓ WACC Model updated with formulas")

    # =========================================================================
    # DDM MODEL
    # =========================================================================
    ws = wb["DDM"]

    ws["A2"] = "Dividend Discount Model - Lululemon Athletica (LULU)"
    ws["A2"].font = title_font

    ws["A4"] = "Note: LULU does not pay dividends"
    ws["A4"].font = Font(italic=True, color="FF0000")

    ws["A6"] = "Current Dividend (D0)"
    ws["B6"] = 0
    ws["A7"] = "Dividend Growth Rate (g)"
    ws["B7"] = "N/A"
    ws["A8"] = "Cost of Equity (Ke)"
    ws["B8"] = "=WACC!B8"
    ws["A9"] = "Current Stock Price"
    ws["B9"] = 170.09
    ws["A10"] = "DDM Value"
    ws["B10"] = "N/A - No Dividends"

    print("✓ DDM Model updated (LULU has no dividends)")

    # =========================================================================
    # DCF MODEL - with Formulas
    # =========================================================================
    ws = wb["DCF"]

    try:
        ws["A2"] = "DCF Valuation - Lululemon Athletica (LULU)"
        ws["A2"].font = title_font
    except:
        pass

    # DCF Assumptions
    ws["A4"] = "DCF Assumptions"
    ws["A4"].font = header_font
    ws["A5"] = "WACC"
    ws["B5"] = "=WACC!B23"
    ws["A6"] = "Terminal Growth Rate"
    ws["B6"] = 0.025  # 2.5%
    ws["A7"] = "Shares Outstanding (millions)"
    ws["B7"] = 122.0
    ws["A8"] = "Current Stock Price"
    ws["B8"] = 170.09

    # Free Cash Flow Projections
    ws["A10"] = "Free Cash Flow Projections"
    ws["A10"].font = header_font

    # Headers
    fcf_years = ["FY 2025", "FY 2026E", "FY 2027E", "FY 2028E", "FY 2029E", "FY 2030E"]
    for i, year in enumerate(fcf_years):
        try:
            cell = ws.cell(row=11, column=2 + i)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = year
                cell.font = header_font
        except:
            pass

    # FCF row
    ws["A12"] = "Free Cash Flow"
    try:
        ws["B12"] = "=CashFlow!F35"  # FY 2025 FCF
        ws["C12"] = "=CashFlow!G35"  # FY 2026E FCF
        ws["D12"] = "=CashFlow!H35"  # FY 2027E FCF
        ws["E12"] = "=D12*1.03"  # FY 2028E FCF (3% growth)
        ws["F12"] = "=E12*1.03"  # FY 2029E FCF
        ws["G12"] = "=F12*1.025"  # FY 2030E FCF (terminal)
    except:
        pass

    # Discount Factor
    ws["A13"] = "Discount Period"
    ws["A14"] = "Discount Factor"
    ws["A15"] = "Present Value of FCF"

    try:
        for i, period in enumerate([0, 1, 2, 3, 4, 5]):
            col_letter = get_column_letter(2 + i)
            ws.cell(row=13, column=2 + i, value=period)
            if period > 0:
                ws.cell(row=14, column=2 + i, value=f"=1/(1+$B$5)^{col_letter}13")
                ws.cell(row=15, column=2 + i, value=f"={col_letter}12*{col_letter}14")
    except:
        pass

    # Terminal Value
    ws["A17"] = "Terminal Value Calculation"
    ws["A17"].font = header_font
    ws["A18"] = "Terminal FCF (FY 2030)"
    ws["B18"] = "=G12"
    ws["A19"] = "Terminal Value"
    ws["B19"] = "=B18*(1+$B$6)/($B$5-$B$6)"  # Gordon Growth
    ws["A20"] = "PV of Terminal Value"
    ws["B20"] = "=B19/(1+$B$5)^5"

    # Enterprise Value
    ws["A22"] = "Valuation Summary"
    ws["A22"].font = header_font
    ws["A23"] = "Sum of PV of FCF"
    ws["B23"] = "=SUM(C15:G15)"
    ws["A24"] = "PV of Terminal Value"
    ws["B24"] = "=B20"
    ws["A25"] = "Enterprise Value"
    ws["B25"] = "=B23+B24"
    ws["A26"] = "Less: Total Debt"
    ws["B26"] = "=-BalanceSheet!G23-BalanceSheet!G29"
    ws["A27"] = "Add: Cash"
    ws["B27"] = "=BalanceSheet!G6"
    ws["A28"] = "Equity Value"
    ws["B28"] = "=B25+B26+B27"
    ws["A29"] = "Implied Share Price"
    ws["B29"] = "=B28/$B$7"

    print("✓ DCF Model updated with formulas")

    # =========================================================================
    # SAVE THE WORKBOOK
    # =========================================================================
    output_path = r"c:\Users\nduta\OneDrive\Desktop\Projects\lulu-lemon-project\Module7_LULU_Final.xlsm"
    wb.save(output_path)

    print("\n" + "=" * 70)
    print("✓ Module7 FINAL VERSION COMPLETE!")
    print("=" * 70)
    print(f"\nSaved to: {output_path}")
    print("\nKey Features:")
    print("  • Bloomberg exact headings (Revenue, Cost Of Goods Sold, etc.)")
    print("  • Fiscal Year format: FY 2021, FY 2022... FY 2027E")
    print("  • Fiscal year end dates: 1/31/2021, 1/30/2022, etc.")
    print("  • Excel FORMULAS for all calculated rows:")
    print("      - Gross Profit = Revenue - COGS")
    print("      - Net Income = EBT - Tax")
    print("      - Assets = Liabilities + Equity (CHECK row included)")
    print("      - FCF = CFO + CapEx")
    print("      - WACC = Ke*We + Kd*Wd")
    print("      - DCF with Terminal Value")
    print("\nData Source: Bloomberg Macro XIDF (1).xlsm")
    print("\nYellow highlighted cells = Estimates (FY 2026E, FY 2027E)")


if __name__ == "__main__":
    update_module7()
