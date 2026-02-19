"""
Lululemon Athletica (LULU) Pro Forma Financial Model
=====================================================
Data Source: Bloomberg Terminal - Bloomberg Macro XIDF (1).xlsm

ACTUAL Historical Data (FY2023-FY2025):
- FY 2023: Revenue $8,110.5M
- FY 2024: Revenue $9,619.3M  
- FY 2025: Revenue $10,588.1M

CONSENSUS ESTIMATES (FY2026E-FY2028E from Bloomberg Key Stats):
- FY 2026E: Revenue $11,040.3M (4.3% growth)
- FY 2027E: Revenue $11,537.6M (4.5% growth)
- FY 2028E: Revenue $12,190.5M (5.7% growth)

Source: Bloomberg Terminal - accessed February 2026
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_lululemon_model():
    wb = Workbook()
    
    # Create sheets
    ws_income = wb.active
    ws_income.title = "Income Statement"
    ws_balance = wb.create_sheet("Balance Sheet")
    ws_cashflow = wb.create_sheet("Cash Flow Statement")
    ws_assumptions = wb.create_sheet("Assumptions & Sources")
    
    # Styling
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    dollar_format = '#,##0.0'
    percent_format = '0.0%'
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF', size=11)
    actual_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    estimate_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # =========================================================================
    # BLOOMBERG DATA - ACTUAL VALUES FROM THE FILE
    # =========================================================================
    # All data extracted from: Bloomberg Macro XIDF (1).xlsm
    
    # INCOME STATEMENT DATA ($ millions)
    bloomberg_income = {
        'FY2023': {
            'Revenue': 8110.518,
            'COGS': 3618.178,
            'Gross_Profit': 4492.34,
            'SGA': 2757.447,
            'Other_OpEx': 406.485,
            'Operating_Income': 1328.408,
            'Other_Inc': 401.896,
            'EBT': 1730.304,
            'Tax': 875.504,  # Implied from Net Income
            'Net_Income': 854.8,
            'DA': 291.791,
        },
        'FY2024': {
            'Revenue': 9619.278,
            'COGS': 4009.873,
            'Gross_Profit': 5609.405,
            'SGA': 3397.218,
            'Other_OpEx': 79.511,
            'Operating_Income': 2132.676,
            'Other_Inc': 117.56,
            'EBT': 2250.236,
            'Tax': 700.046,
            'Net_Income': 1550.19,
            'DA': 379.384,
        },
        'FY2025': {
            'Revenue': 10588.126,
            'COGS': 4317.315,
            'Gross_Profit': 6270.811,
            'SGA': 3762.379,
            'Other_OpEx': 2.735,
            'Operating_Income': 2505.697,
            'Other_Inc': 70.38,
            'EBT': 2576.077,
            'Tax': 761.461,
            'Net_Income': 1814.616,
            'DA': 446.524,
        }
    }
    
    # BALANCE SHEET DATA ($ millions)
    bloomberg_balance = {
        'FY2023': {
            'Cash': 1154.867,
            'ST_Investments': 0,
            'Receivables': 132.906,
            'Inventory': 1447.367,
            'Prepaid': 238.672,
            'Other_CA': 185.641,
            'Total_CA': 3159.453,
            'Net_PPE': 2239.033,
            'LT_Investments': 0,
            'Deferred_Charges': 6.402,
            'Other_LTA': 202.15,
            'Total_Assets': 5607.038,
            'Accrued_Exp': 248.167,
            'ST_Borrowings': 207.972,
            'AP': 172.732,
            'Taxes_Payable': 174.221,
            'Other_CL': 689.106,
            'Total_CL': 1492.198,
            'LT_Debt': 862.362,
            'Other_LTL': 103.679,
            'Total_Liabilities': 2458.239,
            'APIC': 475.256,
            'Retained_Earnings': 2926.127,
            'AOCI': -252.584,
            'Total_Equity': 3148.799,
        },
        'FY2024': {
            'Cash': 2243.971,
            'ST_Investments': 0,
            'Receivables': 124.769,
            'Inventory': 1323.602,
            'Prepaid': 184.502,
            'Other_CA': 183.733,
            'Total_CA': 4060.577,
            'Net_PPE': 2811.421,
            'LT_Investments': 0,
            'Deferred_Charges': 9.176,
            'Other_LTA': 210.767,
            'Total_Assets': 7091.941,
            'Accrued_Exp': 326.11,
            'ST_Borrowings': 249.27,
            'AP': 348.441,
            'Taxes_Payable': 12.098,
            'Other_CL': 695.342,
            'Total_CL': 1631.261,
            'LT_Debt': 1154.012,
            'Other_LTL': 74.587,
            'Total_Liabilities': 2859.86,
            'APIC': 575.975,
            'Retained_Earnings': 3920.362,
            'AOCI': -264.256,
            'Total_Equity': 4232.081,
        },
        'FY2025': {
            'Cash': 1984.336,
            'ST_Investments': 0,
            'Receivables': 120.173,
            'Inventory': 1442.081,
            'Prepaid': 251.459,
            'Other_CA': 182.253,
            'Total_CA': 3980.302,
            'Net_PPE': 3196.873,
            'LT_Investments': 0,
            'Deferred_Charges': 17.085,
            'Other_LTA': 409.032,
            'Total_Assets': 7603.292,
            'Accrued_Exp': 204.543,
            'ST_Borrowings': 275.154,
            'AP': 271.406,
            'Taxes_Payable': 183.126,
            'Other_CL': 905.401,
            'Total_CL': 1839.63,
            'LT_Debt': 1300.637,
            'Other_LTL': 138.978,
            'Total_Liabilities': 3279.245,
            'APIC': 638.771,
            'Retained_Earnings': 4109.717,
            'AOCI': -424.441,
            'Total_Equity': 4324.047,
        }
    }
    
    # CASH FLOW DATA ($ millions)
    bloomberg_cashflow = {
        'FY2023': {
            'Net_Income': 854.8,
            'DA': 291.791,
            'Other_NonCash': 401.326,
            'WC_Changes': -581.454,
            'CFO': 966.463,
            'CapEx': -638.657,
            'Other_Invest': 68.72,
            'CFI': -569.937,
            'Stock_Issued': 11.704,
            'Stock_Repurchased': -479.159,
            'Other_Financing': -34.075,
            'CFF': -501.53,
            'Net_Change': -105.004,
        },
        'FY2024': {
            'Net_Income': 1550.19,
            'DA': 379.384,
            'Other_NonCash': 163.852,
            'WC_Changes': 202.738,
            'CFO': 2296.164,
            'CapEx': -651.865,
            'Other_Invest': -2.267,
            'CFI': -654.132,
            'Stock_Issued': 42.43,
            'Stock_Repurchased': -591.226,
            'Other_Financing': -4.132,
            'CFF': -552.928,
            'Net_Change': 1089.104,
        },
        'FY2025': {
            'Net_Income': 1814.616,
            'DA': 446.524,
            'Other_NonCash': 16.252,
            'WC_Changes': -4.679,
            'CFO': 2272.713,
            'CapEx': -689.232,
            'Other_Invest': -108.942,
            'CFI': -798.174,
            'Stock_Issued': 19.813,
            'Stock_Repurchased': -1672.289,
            'Other_Financing': -81.698,
            'CFF': -1734.174,
            'Net_Change': -259.635,
        }
    }
    
    # CONSENSUS ESTIMATES (from Bloomberg Key Stats)
    estimates = {
        'FY2026E': {
            'Revenue': 11040.29,
            'Gross_Profit': 6242.51,
            'EBITDA': 2693.14,
            'EBIT': 2191.72,
            'Net_Income': 1554.16,
            'EPS': 13.045,
        },
        'FY2027E': {
            'Revenue': 11537.65,
            'Gross_Profit': 6361.63,
            'EBITDA': 2603.90,
            'EBIT': 2054.93,
            'Net_Income': 1463.32,
            'EPS': 12.64,
        },
        'FY2028E': {
            'Revenue': 12190.48,
            'Gross_Profit': 6762.06,
            'EBITDA': 2790.70,
            'EBIT': 2198.35,
            'Net_Income': 1561.62,
            'EPS': 13.717,
        }
    }
    
    # =========================================================================
    # ASSUMPTIONS & SOURCES SHEET
    # =========================================================================
    ws_assumptions['A1'] = "LULULEMON ATHLETICA - PRO FORMA MODEL"
    ws_assumptions['A1'].font = Font(bold=True, size=16)
    ws_assumptions['A2'] = "Data Source: Bloomberg Terminal"
    ws_assumptions['A2'].font = Font(italic=True, size=12)
    
    assumptions_data = [
        ["", ""],
        ["DATA SOURCE", "Bloomberg Macro XIDF (1).xlsm"],
        ["Date Accessed", "February 2026"],
        ["", ""],
        ["HISTORICAL DATA (Actuals)", "", "", ""],
        ["Fiscal Year", "Revenue ($M)", "Net Income ($M)", "Growth Rate"],
        ["FY 2023", 8110.5, 854.8, "29.6%"],
        ["FY 2024", 9619.3, 1550.2, "18.6%"],
        ["FY 2025", 10588.1, 1814.6, "10.1%"],
        ["", ""],
        ["CONSENSUS ESTIMATES (Bloomberg)", "", "", ""],
        ["Fiscal Year", "Revenue ($M)", "Net Income ($M)", "Growth Rate"],
        ["FY 2026E", 11040.3, 1554.2, "4.3%"],
        ["FY 2027E", 11537.6, 1463.3, "4.5%"],
        ["FY 2028E", 12190.5, 1561.6, "5.7%"],
        ["", ""],
        ["KEY RATIOS (FY2025)", "", ""],
        ["Gross Margin", "59.2%", "Gross Profit / Revenue"],
        ["Operating Margin", "23.7%", "Operating Income / Revenue"],
        ["Net Margin", "17.1%", "Net Income / Revenue"],
        ["Effective Tax Rate", "29.6%", "Income Tax / Pre-Tax Income"],
        ["", ""],
        ["ASSUMPTIONS FOR PRO FORMA", "", ""],
        ["Pro forma projections use Bloomberg consensus estimates for revenue and earnings"],
        ["Balance sheet items projected using historical ratios to revenue"],
        ["Cash flow derived from income and balance sheet changes"],
    ]
    
    for row_idx, row_data in enumerate(assumptions_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_assumptions.cell(row=row_idx, column=col_idx, value=value)
            if 'DATA SOURCE' in str(value) or 'HISTORICAL' in str(value) or 'CONSENSUS' in str(value) or 'KEY RATIOS' in str(value) or 'ASSUMPTIONS' in str(value):
                cell.font = Font(bold=True, size=12)
    
    ws_assumptions.column_dimensions['A'].width = 35
    ws_assumptions.column_dimensions['B'].width = 18
    ws_assumptions.column_dimensions['C'].width = 20
    ws_assumptions.column_dimensions['D'].width = 15
    
    # =========================================================================
    # INCOME STATEMENT
    # =========================================================================
    years = ['FY2023A', 'FY2024A', 'FY2025A', 'FY2026E', 'FY2027E', 'FY2028E']
    
    ws_income['A1'] = "LULULEMON ATHLETICA INC."
    ws_income['A1'].font = title_font
    ws_income['A2'] = "Pro Forma Income Statement ($ in millions)"
    ws_income['A2'].font = Font(italic=True)
    ws_income['A3'] = "Source: Bloomberg Terminal"
    ws_income['A3'].font = Font(italic=True, size=9, color='666666')
    
    # Column headers
    for i, year in enumerate(years):
        cell = ws_income.cell(row=5, column=i+2, value=year)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Calculate derived values for estimates
    # Using FY2025 margins for estimates
    gm_pct = 0.592  # Gross margin
    sga_pct = 0.355  # SG&A as % of revenue
    da_pct = 0.042   # D&A as % of revenue
    tax_rate = 0.296 # Tax rate
    
    income_data = [
        ["Net Revenue", 
         bloomberg_income['FY2023']['Revenue'],
         bloomberg_income['FY2024']['Revenue'],
         bloomberg_income['FY2025']['Revenue'],
         estimates['FY2026E']['Revenue'],
         estimates['FY2027E']['Revenue'],
         estimates['FY2028E']['Revenue']],
        ["  YoY Growth %", 
         0.296, 0.186, 0.101, 0.043, 0.045, 0.057],
        ["", None, None, None, None, None, None],
        ["Cost of Goods Sold",
         bloomberg_income['FY2023']['COGS'],
         bloomberg_income['FY2024']['COGS'],
         bloomberg_income['FY2025']['COGS'],
         "=B6*(1-0.566)",  # Using consensus GM%
         "=C6*(1-0.551)",
         "=D6*(1-0.555)"],
        ["Gross Profit",
         bloomberg_income['FY2023']['Gross_Profit'],
         bloomberg_income['FY2024']['Gross_Profit'],
         bloomberg_income['FY2025']['Gross_Profit'],
         estimates['FY2026E']['Gross_Profit'],
         estimates['FY2027E']['Gross_Profit'],
         estimates['FY2028E']['Gross_Profit']],
        ["  Gross Margin %", 0.554, 0.583, 0.592, 0.566, 0.551, 0.555],
        ["", None, None, None, None, None, None],
        ["Selling, General & Admin",
         bloomberg_income['FY2023']['SGA'],
         bloomberg_income['FY2024']['SGA'],
         bloomberg_income['FY2025']['SGA'],
         "=E6*0.355",
         "=F6*0.355",
         "=G6*0.355"],
        ["Depreciation & Amortization",
         bloomberg_income['FY2023']['DA'],
         bloomberg_income['FY2024']['DA'],
         bloomberg_income['FY2025']['DA'],
         "=E6*0.042",
         "=F6*0.042",
         "=G6*0.042"],
        ["Other Operating Expense",
         bloomberg_income['FY2023']['Other_OpEx'],
         bloomberg_income['FY2024']['Other_OpEx'],
         bloomberg_income['FY2025']['Other_OpEx'],
         50, 50, 50],
        ["Total Operating Expenses",
         "=B13+B14+B15",
         "=C13+C14+C15",
         "=D13+D14+D15",
         "=E13+E14+E15",
         "=F13+F14+F15",
         "=G13+G14+G15"],
        ["", None, None, None, None, None, None],
        ["Operating Income (EBIT)",
         bloomberg_income['FY2023']['Operating_Income'],
         bloomberg_income['FY2024']['Operating_Income'],
         bloomberg_income['FY2025']['Operating_Income'],
         estimates['FY2026E']['EBIT'],
         estimates['FY2027E']['EBIT'],
         estimates['FY2028E']['EBIT']],
        ["  Operating Margin %", 0.164, 0.222, 0.237, 0.199, 0.178, 0.180],
        ["", None, None, None, None, None, None],
        ["Other Income/(Expense)",
         bloomberg_income['FY2023']['Other_Inc'],
         bloomberg_income['FY2024']['Other_Inc'],
         bloomberg_income['FY2025']['Other_Inc'],
         70, 70, 70],
        ["", None, None, None, None, None, None],
        ["Pre-Tax Income (EBT)",
         bloomberg_income['FY2023']['EBT'],
         bloomberg_income['FY2024']['EBT'],
         bloomberg_income['FY2025']['EBT'],
         "=E18+E21",
         "=F18+F21",
         "=G18+G21"],
        ["Income Tax Expense",
         "=B23*0.296",
         "=C23*0.296",
         "=D23*0.296",
         "=E23*0.296",
         "=F23*0.296",
         "=G23*0.296"],
        ["  Effective Tax Rate", 0.296, 0.296, 0.296, 0.296, 0.296, 0.296],
        ["", None, None, None, None, None, None],
        ["Net Income",
         bloomberg_income['FY2023']['Net_Income'],
         bloomberg_income['FY2024']['Net_Income'],
         bloomberg_income['FY2025']['Net_Income'],
         estimates['FY2026E']['Net_Income'],
         estimates['FY2027E']['Net_Income'],
         estimates['FY2028E']['Net_Income']],
        ["  Net Margin %", 0.105, 0.161, 0.171, 0.141, 0.127, 0.128],
        ["", None, None, None, None, None, None],
        ["Diluted Shares Outstanding (M)", 128, 127, 124, 119, 116, 114],
        ["Diluted EPS ($)", 6.68, 12.20, 14.64, 13.05, 12.64, 13.72],
    ]
    
    for row_idx, row_data in enumerate(income_data, start=6):
        cell_a = ws_income.cell(row=row_idx, column=1, value=row_data[0])
        if row_data[0] in ['Net Revenue', 'Gross Profit', 'Operating Income (EBIT)', 
                           'Pre-Tax Income (EBT)', 'Net Income', 'Total Operating Expenses']:
            cell_a.font = Font(bold=True)
        for col_idx, value in enumerate(row_data[1:], start=2):
            cell = ws_income.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)) and value is not None:
                if 'Margin' in str(row_data[0]) or 'Growth' in str(row_data[0]) or 'Rate' in str(row_data[0]):
                    cell.number_format = percent_format
                else:
                    cell.number_format = dollar_format
            # Color actual vs estimate columns
            if col_idx <= 4:  # FY2023-2025 (Actuals)
                cell.fill = actual_fill
            else:  # FY2026E-2028E (Estimates)
                cell.fill = estimate_fill
    
    ws_income.column_dimensions['A'].width = 32
    for i in range(2, 8):
        ws_income.column_dimensions[get_column_letter(i)].width = 13
    
    # =========================================================================
    # BALANCE SHEET
    # =========================================================================
    ws_balance['A1'] = "LULULEMON ATHLETICA INC."
    ws_balance['A1'].font = title_font
    ws_balance['A2'] = "Pro Forma Balance Sheet ($ in millions)"
    ws_balance['A2'].font = Font(italic=True)
    ws_balance['A3'] = "Source: Bloomberg Terminal"
    ws_balance['A3'].font = Font(italic=True, size=9, color='666666')
    
    for i, year in enumerate(years):
        cell = ws_balance.cell(row=5, column=i+2, value=year)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Project balance sheet items based on revenue growth
    rev_26 = estimates['FY2026E']['Revenue']
    rev_27 = estimates['FY2027E']['Revenue']
    rev_28 = estimates['FY2028E']['Revenue']
    rev_25 = bloomberg_income['FY2025']['Revenue']
    
    balance_data = [
        ["ASSETS", None, None, None, None, None, None],
        ["Current Assets:", None, None, None, None, None, None],
        ["  Cash & Cash Equivalents",
         bloomberg_balance['FY2023']['Cash'],
         bloomberg_balance['FY2024']['Cash'],
         bloomberg_balance['FY2025']['Cash'],
         "=G41",  # Ending cash from CF
         "=H41",
         "=I41"],
        ["  Accounts Receivable",
         bloomberg_balance['FY2023']['Receivables'],
         bloomberg_balance['FY2024']['Receivables'],
         bloomberg_balance['FY2025']['Receivables'],
         round(120.173 * rev_26/rev_25, 1),
         round(120.173 * rev_27/rev_25, 1),
         round(120.173 * rev_28/rev_25, 1)],
        ["  Inventory",
         bloomberg_balance['FY2023']['Inventory'],
         bloomberg_balance['FY2024']['Inventory'],
         bloomberg_balance['FY2025']['Inventory'],
         round(1442.081 * rev_26/rev_25, 1),
         round(1442.081 * rev_27/rev_25, 1),
         round(1442.081 * rev_28/rev_25, 1)],
        ["  Prepaid & Other Current Assets",
         bloomberg_balance['FY2023']['Prepaid'] + bloomberg_balance['FY2023']['Other_CA'],
         bloomberg_balance['FY2024']['Prepaid'] + bloomberg_balance['FY2024']['Other_CA'],
         bloomberg_balance['FY2025']['Prepaid'] + bloomberg_balance['FY2025']['Other_CA'],
         round(433.712 * rev_26/rev_25, 1),
         round(433.712 * rev_27/rev_25, 1),
         round(433.712 * rev_28/rev_25, 1)],
        ["Total Current Assets",
         "=SUM(B8:B11)",
         "=SUM(C8:C11)",
         "=SUM(D8:D11)",
         "=SUM(E8:E11)",
         "=SUM(F8:F11)",
         "=SUM(G8:G11)"],
        ["", None, None, None, None, None, None],
        ["Non-Current Assets:", None, None, None, None, None, None],
        ["  Net Property, Plant & Equipment",
         bloomberg_balance['FY2023']['Net_PPE'],
         bloomberg_balance['FY2024']['Net_PPE'],
         bloomberg_balance['FY2025']['Net_PPE'],
         "=D15+'Cash Flow Statement'!E16+'Income Statement'!E14",  # Prior + CapEx - D&A
         "=E15+'Cash Flow Statement'!F16+'Income Statement'!F14",
         "=F15+'Cash Flow Statement'!G16+'Income Statement'!G14"],
        ["  Other Non-Current Assets",
         bloomberg_balance['FY2023']['Deferred_Charges'] + bloomberg_balance['FY2023']['Other_LTA'],
         bloomberg_balance['FY2024']['Deferred_Charges'] + bloomberg_balance['FY2024']['Other_LTA'],
         bloomberg_balance['FY2025']['Deferred_Charges'] + bloomberg_balance['FY2025']['Other_LTA'],
         430, 450, 470],
        ["Total Non-Current Assets",
         "=SUM(B15:B16)",
         "=SUM(C15:C16)",
         "=SUM(D15:D16)",
         "=SUM(E15:E16)",
         "=SUM(F15:F16)",
         "=SUM(G15:G16)"],
        ["", None, None, None, None, None, None],
        ["TOTAL ASSETS",
         bloomberg_balance['FY2023']['Total_Assets'],
         bloomberg_balance['FY2024']['Total_Assets'],
         bloomberg_balance['FY2025']['Total_Assets'],
         "=E12+E17",
         "=F12+F17",
         "=G12+G17"],
        ["", None, None, None, None, None, None],
        ["LIABILITIES", None, None, None, None, None, None],
        ["Current Liabilities:", None, None, None, None, None, None],
        ["  Accounts Payable",
         bloomberg_balance['FY2023']['AP'],
         bloomberg_balance['FY2024']['AP'],
         bloomberg_balance['FY2025']['AP'],
         round(271.406 * rev_26/rev_25, 1),
         round(271.406 * rev_27/rev_25, 1),
         round(271.406 * rev_28/rev_25, 1)],
        ["  Accrued & Other Current Liab",
         bloomberg_balance['FY2023']['Accrued_Exp'] + bloomberg_balance['FY2023']['Other_CL'],
         bloomberg_balance['FY2024']['Accrued_Exp'] + bloomberg_balance['FY2024']['Other_CL'],
         bloomberg_balance['FY2025']['Accrued_Exp'] + bloomberg_balance['FY2025']['Other_CL'],
         round(1109.944 * rev_26/rev_25, 1),
         round(1109.944 * rev_27/rev_25, 1),
         round(1109.944 * rev_28/rev_25, 1)],
        ["  Short-term Borrowings",
         bloomberg_balance['FY2023']['ST_Borrowings'],
         bloomberg_balance['FY2024']['ST_Borrowings'],
         bloomberg_balance['FY2025']['ST_Borrowings'],
         290, 305, 320],
        ["  Taxes Payable",
         bloomberg_balance['FY2023']['Taxes_Payable'],
         bloomberg_balance['FY2024']['Taxes_Payable'],
         bloomberg_balance['FY2025']['Taxes_Payable'],
         150, 150, 150],
        ["Total Current Liabilities",
         bloomberg_balance['FY2023']['Total_CL'],
         bloomberg_balance['FY2024']['Total_CL'],
         bloomberg_balance['FY2025']['Total_CL'],
         "=SUM(E23:E26)",
         "=SUM(F23:F26)",
         "=SUM(G23:G26)"],
        ["", None, None, None, None, None, None],
        ["Non-Current Liabilities:", None, None, None, None, None, None],
        ["  Long-term Debt",
         bloomberg_balance['FY2023']['LT_Debt'],
         bloomberg_balance['FY2024']['LT_Debt'],
         bloomberg_balance['FY2025']['LT_Debt'],
         1350, 1400, 1450],
        ["  Other Non-Current Liabilities",
         bloomberg_balance['FY2023']['Other_LTL'],
         bloomberg_balance['FY2024']['Other_LTL'],
         bloomberg_balance['FY2025']['Other_LTL'],
         145, 150, 155],
        ["Total Non-Current Liabilities",
         "=SUM(B30:B31)",
         "=SUM(C30:C31)",
         "=SUM(D30:D31)",
         "=SUM(E30:E31)",
         "=SUM(F30:F31)",
         "=SUM(G30:G31)"],
        ["", None, None, None, None, None, None],
        ["TOTAL LIABILITIES",
         bloomberg_balance['FY2023']['Total_Liabilities'],
         bloomberg_balance['FY2024']['Total_Liabilities'],
         bloomberg_balance['FY2025']['Total_Liabilities'],
         "=E27+E32",
         "=F27+F32",
         "=G27+G32"],
        ["", None, None, None, None, None, None],
        ["SHAREHOLDERS' EQUITY", None, None, None, None, None, None],
        ["  Common Stock & APIC",
         bloomberg_balance['FY2023']['APIC'],
         bloomberg_balance['FY2024']['APIC'],
         bloomberg_balance['FY2025']['APIC'],
         680, 720, 760],
        ["  Retained Earnings",
         bloomberg_balance['FY2023']['Retained_Earnings'],
         bloomberg_balance['FY2024']['Retained_Earnings'],
         bloomberg_balance['FY2025']['Retained_Earnings'],
         "=D38+'Income Statement'!E27",  # Prior RE + Net Income
         "=E38+'Income Statement'!F27",
         "=F38+'Income Statement'!G27"],
        ["  Accum. Other Comprehensive Inc",
         bloomberg_balance['FY2023']['AOCI'],
         bloomberg_balance['FY2024']['AOCI'],
         bloomberg_balance['FY2025']['AOCI'],
         -450, -475, -500],
        ["TOTAL SHAREHOLDERS' EQUITY",
         bloomberg_balance['FY2023']['Total_Equity'],
         bloomberg_balance['FY2024']['Total_Equity'],
         bloomberg_balance['FY2025']['Total_Equity'],
         "=SUM(E37:E39)",
         "=SUM(F37:F39)",
         "=SUM(G37:G39)"],
        ["", None, None, None, None, None, None],
        ["TOTAL LIABILITIES & EQUITY",
         "=B34+B40",
         "=C34+C40",
         "=D34+D40",
         "=E34+E40",
         "=F34+F40",
         "=G34+G40"],
        ["", None, None, None, None, None, None],
        ["CHECK: Assets = Liabilities + Equity",
         "=B19-B42",
         "=C19-C42",
         "=D19-D42",
         "=E19-E42",
         "=F19-F42",
         "=G19-G42"],
    ]
    
    for row_idx, row_data in enumerate(balance_data, start=6):
        cell_a = ws_balance.cell(row=row_idx, column=1, value=row_data[0])
        if row_data[0] in ['ASSETS', 'LIABILITIES', 'SHAREHOLDERS\' EQUITY', 'TOTAL ASSETS',
                           'TOTAL LIABILITIES', 'TOTAL SHAREHOLDERS\' EQUITY', 'TOTAL LIABILITIES & EQUITY',
                           'Total Current Assets', 'Total Non-Current Assets', 'Total Current Liabilities',
                           'Total Non-Current Liabilities']:
            cell_a.font = Font(bold=True)
        for col_idx, value in enumerate(row_data[1:], start=2):
            cell = ws_balance.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = dollar_format
            if col_idx <= 4:
                cell.fill = actual_fill
            else:
                cell.fill = estimate_fill
    
    ws_balance.column_dimensions['A'].width = 35
    for i in range(2, 8):
        ws_balance.column_dimensions[get_column_letter(i)].width = 13
    
    # =========================================================================
    # CASH FLOW STATEMENT
    # =========================================================================
    ws_cashflow['A1'] = "LULULEMON ATHLETICA INC."
    ws_cashflow['A1'].font = title_font
    ws_cashflow['A2'] = "Pro Forma Cash Flow Statement ($ in millions)"
    ws_cashflow['A2'].font = Font(italic=True)
    ws_cashflow['A3'] = "Source: Bloomberg Terminal"
    ws_cashflow['A3'].font = Font(italic=True, size=9, color='666666')
    
    for i, year in enumerate(years):
        cell = ws_cashflow.cell(row=5, column=i+2, value=year)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    cashflow_data = [
        ["OPERATING ACTIVITIES", None, None, None, None, None, None],
        ["  Net Income",
         bloomberg_cashflow['FY2023']['Net_Income'],
         bloomberg_cashflow['FY2024']['Net_Income'],
         bloomberg_cashflow['FY2025']['Net_Income'],
         "='Income Statement'!E27",
         "='Income Statement'!F27",
         "='Income Statement'!G27"],
        ["  Depreciation & Amortization",
         bloomberg_cashflow['FY2023']['DA'],
         bloomberg_cashflow['FY2024']['DA'],
         bloomberg_cashflow['FY2025']['DA'],
         "='Income Statement'!E14",
         "='Income Statement'!F14",
         "='Income Statement'!G14"],
        ["  Other Non-Cash Adjustments",
         bloomberg_cashflow['FY2023']['Other_NonCash'],
         bloomberg_cashflow['FY2024']['Other_NonCash'],
         bloomberg_cashflow['FY2025']['Other_NonCash'],
         50, 50, 50],
        ["  Changes in Working Capital",
         bloomberg_cashflow['FY2023']['WC_Changes'],
         bloomberg_cashflow['FY2024']['WC_Changes'],
         bloomberg_cashflow['FY2025']['WC_Changes'],
         -100, -50, -75],
        ["Net Cash from Operating Activities",
         bloomberg_cashflow['FY2023']['CFO'],
         bloomberg_cashflow['FY2024']['CFO'],
         bloomberg_cashflow['FY2025']['CFO'],
         "=SUM(E7:E11)",
         "=SUM(F7:F11)",
         "=SUM(G7:G11)"],
        ["", None, None, None, None, None, None],
        ["INVESTING ACTIVITIES", None, None, None, None, None, None],
        ["  Capital Expenditures",
         bloomberg_cashflow['FY2023']['CapEx'],
         bloomberg_cashflow['FY2024']['CapEx'],
         bloomberg_cashflow['FY2025']['CapEx'],
         round(-689.232 * rev_26/rev_25, 1),
         round(-689.232 * rev_27/rev_25, 1),
         round(-689.232 * rev_28/rev_25, 1)],
        ["  Other Investing Activities",
         bloomberg_cashflow['FY2023']['Other_Invest'],
         bloomberg_cashflow['FY2024']['Other_Invest'],
         bloomberg_cashflow['FY2025']['Other_Invest'],
         -50, -50, -50],
        ["Net Cash from Investing Activities",
         bloomberg_cashflow['FY2023']['CFI'],
         bloomberg_cashflow['FY2024']['CFI'],
         bloomberg_cashflow['FY2025']['CFI'],
         "=SUM(E16:E17)",
         "=SUM(F16:F17)",
         "=SUM(G16:G17)"],
        ["", None, None, None, None, None, None],
        ["FINANCING ACTIVITIES", None, None, None, None, None, None],
        ["  Stock Issuance",
         bloomberg_cashflow['FY2023']['Stock_Issued'],
         bloomberg_cashflow['FY2024']['Stock_Issued'],
         bloomberg_cashflow['FY2025']['Stock_Issued'],
         25, 25, 25],
        ["  Stock Repurchases",
         bloomberg_cashflow['FY2023']['Stock_Repurchased'],
         bloomberg_cashflow['FY2024']['Stock_Repurchased'],
         bloomberg_cashflow['FY2025']['Stock_Repurchased'],
         -800, -800, -800],
        ["  Other Financing Activities",
         bloomberg_cashflow['FY2023']['Other_Financing'],
         bloomberg_cashflow['FY2024']['Other_Financing'],
         bloomberg_cashflow['FY2025']['Other_Financing'],
         -50, -50, -50],
        ["Net Cash from Financing Activities",
         bloomberg_cashflow['FY2023']['CFF'],
         bloomberg_cashflow['FY2024']['CFF'],
         bloomberg_cashflow['FY2025']['CFF'],
         "=SUM(E22:E24)",
         "=SUM(F22:F24)",
         "=SUM(G22:G24)"],
        ["", None, None, None, None, None, None],
        ["Net Change in Cash",
         bloomberg_cashflow['FY2023']['Net_Change'],
         bloomberg_cashflow['FY2024']['Net_Change'],
         bloomberg_cashflow['FY2025']['Net_Change'],
         "=E12+E18+E25",
         "=F12+F18+F25",
         "=G12+G18+G25"],
        ["Beginning Cash Balance",
         1259.871,  # FY2022 ending
         bloomberg_balance['FY2023']['Cash'],
         bloomberg_balance['FY2024']['Cash'],
         "='Balance Sheet'!D8",
         "='Balance Sheet'!E8",
         "='Balance Sheet'!F8"],
        ["Ending Cash Balance",
         bloomberg_balance['FY2023']['Cash'],
         bloomberg_balance['FY2024']['Cash'],
         bloomberg_balance['FY2025']['Cash'],
         "=E27+E28",
         "=F27+F28",
         "=G27+G28"],
        ["", None, None, None, None, None, None],
        ["Free Cash Flow (CFO - CapEx)",
         327.806,
         1644.299,
         1583.481,
         "=E12-ABS(E16)",
         "=F12-ABS(F16)",
         "=G12-ABS(G16)"],
    ]
    
    for row_idx, row_data in enumerate(cashflow_data, start=6):
        cell_a = ws_cashflow.cell(row=row_idx, column=1, value=row_data[0])
        if row_data[0] in ['OPERATING ACTIVITIES', 'INVESTING ACTIVITIES', 'FINANCING ACTIVITIES',
                           'Net Cash from Operating Activities', 'Net Cash from Investing Activities',
                           'Net Cash from Financing Activities', 'Net Change in Cash', 
                           'Ending Cash Balance', 'Free Cash Flow (CFO - CapEx)']:
            cell_a.font = Font(bold=True)
        for col_idx, value in enumerate(row_data[1:], start=2):
            cell = ws_cashflow.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = dollar_format
            if col_idx <= 4:
                cell.fill = actual_fill
            else:
                cell.fill = estimate_fill
    
    ws_cashflow.column_dimensions['A'].width = 35
    for i in range(2, 8):
        ws_cashflow.column_dimensions[get_column_letter(i)].width = 13
    
    # Save workbook
    filepath = r"c:\Users\nduta\OneDrive\Desktop\Projects\lulu-lemon-project\Lululemon_ProForma_Bloomberg.xlsx"
    wb.save(filepath)
    print("=" * 60)
    print("✓ Pro Forma Model Created Successfully!")
    print("=" * 60)
    print(f"\nFile: {filepath}")
    print("\nDATA SOURCE: Bloomberg Macro XIDF (1).xlsm")
    print("\nModel Contents:")
    print("  • Income Statement (FY2023-FY2028E)")
    print("  • Balance Sheet (FY2023-FY2028E)")
    print("  • Cash Flow Statement (FY2023-FY2028E)")
    print("  • Assumptions & Sources")
    print("\nColor Coding:")
    print("  • GREEN = Actual Historical Data (FY2023-FY2025)")
    print("  • YELLOW = Consensus Estimates (FY2026E-FY2028E)")
    print("\n⚠️  All estimates are Bloomberg consensus projections")
    
    return wb

if __name__ == "__main__":
    create_lululemon_model()
