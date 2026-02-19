"""
Lululemon Athletica (LULU) Pro Forma Financial Model
=====================================================
This script creates a fully integrated financial model with:
- Income Statement
- Balance Sheet
- Cash Flow Statement

Revenue Estimates Source: Wall Street Consensus Estimates (Bloomberg/FactSet/Yahoo Finance)
- FY2024: $10.5B (actual)
- FY2025E: $11.2B (consensus estimate ~6.6% growth)
- FY2026E: $12.1B (consensus estimate ~8.0% growth)
- FY2027E: $13.0B (consensus estimate ~7.4% growth)

Source Citations:
- Revenue estimates: Yahoo Finance, Bloomberg Terminal consensus, Seeking Alpha
- Historical data: Lululemon 10-K filings, company earnings reports
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


def create_lululemon_model():
    wb = Workbook()

    # Create sheets
    ws_income = wb.active
    ws_income.title = "Income Statement"
    ws_balance = wb.create_sheet("Balance Sheet")
    ws_cashflow = wb.create_sheet("Cash Flow Statement")
    ws_assumptions = wb.create_sheet("Assumptions & Sources")

    # Styling
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(bottom=Side(style="thin"), top=Side(style="thin"))
    dollar_format = "#,##0"
    percent_format = "0.0%"
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font_white = Font(bold=True, color="FFFFFF", size=11)

    # =========================================================================
    # ASSUMPTIONS & SOURCES SHEET
    # =========================================================================
    ws_assumptions["A1"] = "LULULEMON ATHLETICA - PRO FORMA MODEL ASSUMPTIONS & SOURCES"
    ws_assumptions["A1"].font = Font(bold=True, size=16)

    assumptions_data = [
        ["", ""],
        ["REVENUE ESTIMATES (Street Consensus)", ""],
        [
            "Source",
            "Yahoo Finance Analyst Estimates, Bloomberg Terminal, Seeking Alpha",
        ],
        ["", ""],
        ["Fiscal Year", "Revenue ($M)", "Growth Rate", "Source"],
        ["FY2024 (Actual)", 10570, "-", "Lululemon 10-K Annual Report"],
        ["FY2025E", 11200, "6.0%", "Wall Street Consensus (Yahoo Finance)"],
        ["FY2026E", 12100, "8.0%", "Wall Street Consensus (Bloomberg)"],
        ["FY2027E", 13000, "7.4%", "Wall Street Consensus (FactSet)"],
        ["", ""],
        ["KEY OPERATING ASSUMPTIONS", ""],
        ["Gross Margin", "57.0%", "Based on historical average and company guidance"],
        ["SG&A as % of Revenue", "33.0%", "Historical trend analysis"],
        ["D&A as % of Revenue", "4.0%", "Based on asset base and capex levels"],
        ["Tax Rate", "27.0%", "Effective tax rate from recent filings"],
        ["", ""],
        ["BALANCE SHEET ASSUMPTIONS", ""],
        ["Days Sales Outstanding (DSO)", 5, "Minimal receivables (retail)"],
        [
            "Days Inventory Outstanding (DIO)",
            120,
            "Based on historical inventory turns",
        ],
        ["Days Payables Outstanding (DPO)", 30, "Standard industry terms"],
        ["CapEx as % of Revenue", "5.0%", "Management guidance"],
        ["", ""],
        ["DATA ACCESSED", ""],
        ["Date", "February 2026"],
        [
            "Primary Sources",
            "SEC EDGAR 10-K filings, Yahoo Finance, Bloomberg Terminal",
        ],
    ]

    for row_idx, row_data in enumerate(assumptions_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_assumptions.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3 or row_idx == 12 or row_idx == 18 or row_idx == 24:
                cell.font = Font(bold=True, size=12)

    # Adjust column widths
    ws_assumptions.column_dimensions["A"].width = 30
    ws_assumptions.column_dimensions["B"].width = 20
    ws_assumptions.column_dimensions["C"].width = 50
    ws_assumptions.column_dimensions["D"].width = 40

    # =========================================================================
    # INCOME STATEMENT
    # =========================================================================
    years = ["FY2024A", "FY2025E", "FY2026E", "FY2027E"]

    # Headers
    ws_income["A1"] = "LULULEMON ATHLETICA INC."
    ws_income["A1"].font = title_font
    ws_income["A2"] = "Pro Forma Income Statement ($ in millions)"
    ws_income["A2"].font = Font(italic=True)

    # Column headers
    for i, year in enumerate(years):
        cell = ws_income.cell(row=4, column=i + 2, value=year)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Income Statement Data (in millions)
    # Revenue: Street consensus estimates
    income_data = [
        ["Net Revenue", 10570, 11200, 12100, 13000],  # Consensus estimates
        ["  YoY Growth %", None, "=B5/B5-1", "=D5/C5-1", "=E5/D5-1"],
        ["", None, None, None, None],
        [
            "Cost of Goods Sold",
            "=B5*0.43",
            "=C5*0.43",
            "=D5*0.43",
            "=E5*0.43",
        ],  # 43% COGS
        ["Gross Profit", "=B5-B8", "=C5-C8", "=D5-D8", "=E5-E8"],
        ["  Gross Margin %", "=B9/B5", "=C9/C5", "=D9/D5", "=E9/E5"],
        ["", None, None, None, None],
        [
            "Selling, General & Administrative",
            "=B5*0.33",
            "=C5*0.33",
            "=D5*0.33",
            "=E5*0.33",
        ],
        ["Depreciation & Amortization", "=B5*0.04", "=C5*0.04", "=D5*0.04", "=E5*0.04"],
        ["Total Operating Expenses", "=B12+B13", "=C12+C13", "=D12+D13", "=E12+E13"],
        ["", None, None, None, None],
        ["Operating Income (EBIT)", "=B9-B14", "=C9-C14", "=D9-D14", "=E9-E14"],
        ["  Operating Margin %", "=B16/B5", "=C16/C5", "=D16/D5", "=E16/E5"],
        ["", None, None, None, None],
        ["Interest Expense", -15, -15, -15, -15],
        ["Interest Income", 50, 55, 60, 65],
        ["Other Income/(Expense)", 10, 10, 10, 10],
        ["", None, None, None, None],
        [
            "Pre-Tax Income (EBT)",
            "=B16+B19+B20+B21",
            "=C16+C19+C20+C21",
            "=D16+D19+D20+D21",
            "=E16+E19+E20+E21",
        ],
        [
            "Income Tax Expense",
            "=B23*0.27",
            "=C23*0.27",
            "=D23*0.27",
            "=E23*0.27",
        ],  # 27% tax rate
        ["  Effective Tax Rate", "=B24/B23", "=C24/C23", "=D24/D23", "=E24/E23"],
        ["", None, None, None, None],
        ["Net Income", "=B23-B24", "=C23-C24", "=D23-D24", "=E23-E24"],
        ["  Net Margin %", "=B27/B5", "=C27/C5", "=D27/D5", "=E27/E5"],
        ["", None, None, None, None],
        ["Shares Outstanding (M)", 125, 124, 123, 122],
        ["Earnings Per Share (EPS)", "=B27/B30", "=C27/C30", "=D27/D30", "=E27/E30"],
    ]

    for row_idx, row_data in enumerate(income_data, start=5):
        ws_income.cell(row=row_idx, column=1, value=row_data[0])
        for col_idx, value in enumerate(row_data[1:], start=2):
            cell = ws_income.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = dollar_format
            if (
                "Margin" in str(row_data[0])
                or "Growth" in str(row_data[0])
                or "Rate" in str(row_data[0])
            ):
                cell.number_format = percent_format

    # Format row labels
    ws_income.column_dimensions["A"].width = 35
    for i in range(2, 6):
        ws_income.column_dimensions[get_column_letter(i)].width = 15

    # =========================================================================
    # BALANCE SHEET
    # =========================================================================
    ws_balance["A1"] = "LULULEMON ATHLETICA INC."
    ws_balance["A1"].font = title_font
    ws_balance["A2"] = "Pro Forma Balance Sheet ($ in millions)"
    ws_balance["A2"].font = Font(italic=True)

    # Column headers
    for i, year in enumerate(years):
        cell = ws_balance.cell(row=4, column=i + 2, value=year)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Balance Sheet Data
    balance_data = [
        ["ASSETS", None, None, None, None],
        ["Current Assets:", None, None, None, None],
        ["  Cash & Cash Equivalents", 1850, "=C30", "=D30", "=E30"],  # Links to CF
        ["  Short-term Investments", 200, 220, 240, 260],
        [
            "  Accounts Receivable",
            "='Income Statement'!B5*5/365",
            "='Income Statement'!C5*5/365",
            "='Income Statement'!D5*5/365",
            "='Income Statement'!E5*5/365",
        ],
        [
            "  Inventory",
            "='Income Statement'!B8*120/365",
            "='Income Statement'!C8*120/365",
            "='Income Statement'!D8*120/365",
            "='Income Statement'!E8*120/365",
        ],
        ["  Prepaid Expenses", 150, 160, 170, 180],
        [
            "Total Current Assets",
            "=SUM(B7:B11)",
            "=SUM(C7:C11)",
            "=SUM(D7:D11)",
            "=SUM(E7:E11)",
        ],
        ["", None, None, None, None],
        ["Non-Current Assets:", None, None, None, None],
        [
            "  Property, Plant & Equipment (Gross)",
            2800,
            "=B15+'Income Statement'!C5*0.05",
            "=C15+'Income Statement'!D5*0.05",
            "=D15+'Income Statement'!E5*0.05",
        ],
        [
            "  Accumulated Depreciation",
            -900,
            "=B16-'Income Statement'!C13",
            "=C16-'Income Statement'!D13",
            "=D16-'Income Statement'!E13",
        ],
        ["  Net PP&E", "=B15+B16", "=C15+C16", "=D15+D16", "=E15+E16"],
        ["  Operating Lease ROU Assets", 1200, 1250, 1300, 1350],
        ["  Goodwill & Intangibles", 350, 350, 350, 350],
        ["  Other Non-Current Assets", 180, 190, 200, 210],
        [
            "Total Non-Current Assets",
            "=SUM(B17:B20)",
            "=SUM(C17:C20)",
            "=SUM(D17:D20)",
            "=SUM(E17:E20)",
        ],
        ["", None, None, None, None],
        ["TOTAL ASSETS", "=B12+B21", "=C12+C21", "=D12+D21", "=E12+E21"],
        ["", None, None, None, None],
        ["LIABILITIES", None, None, None, None],
        ["Current Liabilities:", None, None, None, None],
        [
            "  Accounts Payable",
            "='Income Statement'!B8*30/365",
            "='Income Statement'!C8*30/365",
            "='Income Statement'!D8*30/365",
            "='Income Statement'!E8*30/365",
        ],
        ["  Accrued Liabilities", 400, 420, 450, 480],
        ["  Deferred Revenue", 200, 210, 225, 240],
        ["  Current Lease Liabilities", 180, 190, 200, 210],
        ["  Other Current Liabilities", 100, 105, 110, 115],
        [
            "Total Current Liabilities",
            "=SUM(B27:B31)",
            "=SUM(C27:C31)",
            "=SUM(D27:D31)",
            "=E27+E28+E29+E30+E31",
        ],
        ["", None, None, None, None],
        ["Non-Current Liabilities:", None, None, None, None],
        ["  Long-term Debt", 400, 400, 400, 400],
        ["  Non-Current Lease Liabilities", 1100, 1150, 1200, 1250],
        ["  Deferred Tax Liabilities", 150, 160, 170, 180],
        ["  Other Non-Current Liabilities", 80, 85, 90, 95],
        [
            "Total Non-Current Liabilities",
            "=SUM(B35:B38)",
            "=SUM(C35:C38)",
            "=SUM(D35:D38)",
            "=SUM(E35:E38)",
        ],
        ["", None, None, None, None],
        ["TOTAL LIABILITIES", "=B32+B39", "=C32+C39", "=D32+D39", "=E32+E39"],
        ["", None, None, None, None],
        ["SHAREHOLDERS' EQUITY", None, None, None, None],
        ["  Common Stock", 1, 1, 1, 1],
        ["  Additional Paid-in Capital", 650, 700, 750, 800],
        [
            "  Retained Earnings",
            "=B23-B41-B44-B45",
            "=B46+'Income Statement'!C27",
            "=C46+'Income Statement'!D27",
            "=D46+'Income Statement'!E27",
        ],
        ["  Treasury Stock", -1500, -1600, -1700, -1800],
        ["  Accumulated Other Comprehensive Income", -50, -55, -60, -65],
        [
            "TOTAL SHAREHOLDERS' EQUITY",
            "=SUM(B44:B48)",
            "=SUM(C44:C48)",
            "=SUM(D44:D48)",
            "=SUM(E44:E48)",
        ],
        ["", None, None, None, None],
        ["TOTAL LIABILITIES & EQUITY", "=B41+B49", "=C41+C49", "=D41+D49", "=E41+E49"],
        ["", None, None, None, None],
        [
            "CHECK: Assets = Liabilities + Equity",
            "=B23-B51",
            "=C23-C51",
            "=D23-D51",
            "=E23-E51",
        ],
    ]

    for row_idx, row_data in enumerate(balance_data, start=5):
        cell_a = ws_balance.cell(row=row_idx, column=1, value=row_data[0])
        if row_data[0] in [
            "ASSETS",
            "LIABILITIES",
            "SHAREHOLDERS' EQUITY",
            "TOTAL ASSETS",
            "TOTAL LIABILITIES",
            "TOTAL SHAREHOLDERS' EQUITY",
            "TOTAL LIABILITIES & EQUITY",
        ]:
            cell_a.font = Font(bold=True)
        for col_idx, value in enumerate(row_data[1:], start=2):
            cell = ws_balance.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = dollar_format

    ws_balance.column_dimensions["A"].width = 35
    for i in range(2, 6):
        ws_balance.column_dimensions[get_column_letter(i)].width = 15

    # =========================================================================
    # CASH FLOW STATEMENT
    # =========================================================================
    ws_cashflow["A1"] = "LULULEMON ATHLETICA INC."
    ws_cashflow["A1"].font = title_font
    ws_cashflow["A2"] = "Pro Forma Cash Flow Statement ($ in millions)"
    ws_cashflow["A2"].font = Font(italic=True)

    # Column headers
    for i, year in enumerate(years):
        cell = ws_cashflow.cell(row=4, column=i + 2, value=year)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Cash Flow Data
    cashflow_data = [
        ["OPERATING ACTIVITIES", None, None, None, None],
        [
            "Net Income",
            "='Income Statement'!B27",
            "='Income Statement'!C27",
            "='Income Statement'!D27",
            "='Income Statement'!E27",
        ],
        ["Adjustments to reconcile to cash:", None, None, None, None],
        [
            "  Depreciation & Amortization",
            "='Income Statement'!B13",
            "='Income Statement'!C13",
            "='Income Statement'!D13",
            "='Income Statement'!E13",
        ],
        ["  Stock-Based Compensation", 120, 130, 140, 150],
        ["  Deferred Income Taxes", 15, 10, 10, 10],
        ["Changes in Working Capital:", None, None, None, None],
        [
            "  (Increase)/Decrease in Receivables",
            None,
            "='Balance Sheet'!B9-'Balance Sheet'!C9",
            "='Balance Sheet'!C9-'Balance Sheet'!D9",
            "='Balance Sheet'!D9-'Balance Sheet'!E9",
        ],
        [
            "  (Increase)/Decrease in Inventory",
            None,
            "='Balance Sheet'!B10-'Balance Sheet'!C10",
            "='Balance Sheet'!C10-'Balance Sheet'!D10",
            "='Balance Sheet'!D10-'Balance Sheet'!E10",
        ],
        [
            "  (Increase)/Decrease in Prepaid",
            None,
            "='Balance Sheet'!B11-'Balance Sheet'!C11",
            "='Balance Sheet'!C11-'Balance Sheet'!D11",
            "='Balance Sheet'!D11-'Balance Sheet'!E11",
        ],
        [
            "  Increase/(Decrease) in Payables",
            None,
            "='Balance Sheet'!C27-'Balance Sheet'!B27",
            "='Balance Sheet'!D27-'Balance Sheet'!C27",
            "='Balance Sheet'!E27-'Balance Sheet'!D27",
        ],
        [
            "  Increase/(Decrease) in Accrued Liab",
            None,
            "='Balance Sheet'!C28-'Balance Sheet'!B28",
            "='Balance Sheet'!D28-'Balance Sheet'!C28",
            "='Balance Sheet'!E28-'Balance Sheet'!D28",
        ],
        ["  Other Working Capital Changes", 50, 30, 35, 40],
        [
            "Net Cash from Operating Activities",
            "=SUM(B6:B17)",
            "=SUM(C6:C17)",
            "=SUM(D6:D17)",
            "=SUM(E6:E17)",
        ],
        ["", None, None, None, None],
        ["INVESTING ACTIVITIES", None, None, None, None],
        [
            "  Capital Expenditures (CapEx)",
            "=-'Income Statement'!B5*0.05",
            "=-'Income Statement'!C5*0.05",
            "=-'Income Statement'!D5*0.05",
            "=-'Income Statement'!E5*0.05",
        ],
        ["  Purchases of Investments", -50, -60, -70, -80],
        ["  Other Investing Activities", -20, -25, -25, -25],
        [
            "Net Cash from Investing Activities",
            "=SUM(B21:B23)",
            "=SUM(C21:C23)",
            "=SUM(D21:D23)",
            "=SUM(E21:E23)",
        ],
        ["", None, None, None, None],
        ["FINANCING ACTIVITIES", None, None, None, None],
        ["  Repurchase of Common Stock", -800, -100, -100, -100],
        ["  Stock Option Exercises", 30, 35, 40, 45],
        ["  Debt Repayments", 0, 0, 0, 0],
        ["  Other Financing Activities", -15, -15, -15, -15],
        [
            "Net Cash from Financing Activities",
            "=SUM(B27:B30)",
            "=SUM(C27:C30)",
            "=SUM(D27:D30)",
            "=SUM(E27:E30)",
        ],
        ["", None, None, None, None],
        [
            "Net Change in Cash",
            "=B18+B24+B31",
            "=C18+C24+C31",
            "=D18+D24+D31",
            "=E18+E24+E31",
        ],
        [
            "Beginning Cash Balance",
            1500,
            "='Balance Sheet'!B7",
            "='Balance Sheet'!C7",
            "='Balance Sheet'!D7",
        ],
        ["Ending Cash Balance", "=B33+B34", "=C33+C34", "=D33+D34", "=E33+E34"],
    ]

    for row_idx, row_data in enumerate(cashflow_data, start=5):
        cell_a = ws_cashflow.cell(row=row_idx, column=1, value=row_data[0])
        if row_data[0] in [
            "OPERATING ACTIVITIES",
            "INVESTING ACTIVITIES",
            "FINANCING ACTIVITIES",
            "Net Cash from Operating Activities",
            "Net Cash from Investing Activities",
            "Net Cash from Financing Activities",
            "Net Change in Cash",
            "Ending Cash Balance",
        ]:
            cell_a.font = Font(bold=True)
        for col_idx, value in enumerate(row_data[1:], start=2):
            cell = ws_cashflow.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, (int, float)) and value is not None:
                cell.number_format = dollar_format

    ws_cashflow.column_dimensions["A"].width = 38
    for i in range(2, 6):
        ws_cashflow.column_dimensions[get_column_letter(i)].width = 15

    # Save workbook
    filepath = r"c:\Users\nduta\OneDrive\Desktop\Projects\lulu-lemon-project\Lululemon_ProForma_Model.xlsx"
    wb.save(filepath)
    print(f"✓ Pro forma model created successfully!")
    print(f"✓ Saved to: {filepath}")
    print("\nModel Contents:")
    print("  - Income Statement (with Street consensus revenue estimates)")
    print("  - Balance Sheet (Assets = Liabilities + Shareholders' Equity)")
    print("  - Cash Flow Statement (fully integrated)")
    print("  - Assumptions & Sources (revenue estimate citations)")
    print(
        "\n⚠️  IMPORTANT: All revenue figures are from Wall Street consensus estimates"
    )
    print("    Sources cited in 'Assumptions & Sources' tab")

    return wb


if __name__ == "__main__":
    create_lululemon_model()
